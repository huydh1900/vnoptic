# -*- coding: utf-8 -*-

import base64
import json
import logging
from io import BytesIO
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from ..utils import excel_reader, data_cache, import_validator, excel_template_generator

_logger = logging.getLogger(__name__)


class ProductExcelImport(models.TransientModel):
    _name = 'product.excel.import'
    _description = 'Import Products from Excel'
    
    # File upload
    excel_file = fields.Binary(
        string='Chọn file Excel',
        help='Upload Excel file with product data'
    )
    file_name = fields.Char('File Name')
    
    # Auto-detected product type
    product_type = fields.Selection([
        ('lens', 'Lens'),
        ('opt', 'Optical Product'),
        ('accessory', 'Accessory')
    ], string='Loại sản phẩm', readonly=True)
    
    # Preview data (One2many for table view with pagination)
    preview_line_ids = fields.One2many(
        'product.excel.preview.line',
        'wizard_id',
        string='Preview Lines',
        readonly=True
    )
    
    # Keep JSON for debugging
    preview_data = fields.Text(
        string='Preview Data',
        readonly=True,
        help='Parsed data from Excel file in JSON format'
    )
    
    # Wizard state
    state = fields.Selection([
        ('upload', 'Upload File'),
        ('preview', 'Preview Data'),
        ('done', 'Import Complete')
    ], string='State', default='upload')
    
    # Results
    success_count = fields.Integer('Sản phẩm thành công', readonly=True, default=0)
    error_count = fields.Integer('Số lỗi', readonly=True, default=0)
    error_log = fields.Text('Error Log', readonly=True)
    
    # Template download selection
    template_type = fields.Selection([
        ('lens', 'Mẫu Mắt (Lens)'),
        ('opt', 'Mẫu Gọng/Kính'),
        ('accessory', 'Mẫu Phụ kiện'),
    ], string='Tải mẫu nhập liệu')
    
    # Employee assignment
    employee_id = fields.Many2one(
        'hr.employee',
        string='Nhân viên phụ trách',
        help='Assign all imported products to this employee'
    )
    
    @api.onchange('template_type')
    def _onchange_template_type(self):
        """Auto download template when selection changes"""
        if self.template_type:
            # Reset selection after download
            template_type = self.template_type
            self.template_type = False
            
            try:
                if template_type == 'lens':
                    template_data = self._get_import_template_data('lens')
                    filename = f"BangNhap_Mat_{datetime.now().strftime('%Y%m%d')}.xlsx"
                elif template_type == 'opt':
                    template_data = self._get_import_template_data('opt')
                    filename = f"BangNhap_Gong_{datetime.now().strftime('%Y%m%d')}.xlsx"
                elif template_type == 'accessory':
                    template_data = self._get_import_template_data('accessory')
                    filename = f"BangNhap_PhuKien_{datetime.now().strftime('%Y%m%d')}.xlsx"
                else:
                    return
                
                # Create attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename,
                    'type': 'binary',
                    'datas': base64.b64encode(template_data),
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })
                
                # Return download action
                return {
                    'type': 'ir.actions.act_url',
                    'url': f'/web/content/{attachment.id}?download=true',
                    'target': 'new',
                }
            except Exception as e:
                _logger.error(f"Error generating template: {e}", exc_info=True)
                return {
                    'warning': {
                        'title': 'Lỗi',
                        'message': f'Không thể tạo mẫu: {str(e)}'
                    }
                }
    
    # ===========================================
    # DOWNLOAD TEMPLATE ACTIONS
    # ===========================================
    
    def action_download_lens_template(self):
        """Download Excel template for Lens products"""
        self.ensure_one()
        try:
            template_data = self._get_import_template_data('lens')
            filename = f"BangNhap_Mat_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return self._create_download_action(template_data, filename)
        except Exception as e:
            _logger.error(f"Error generating lens template: {e}", exc_info=True)
            raise UserError(_('Error generating template: %s') % str(e))
    
    def action_download_opt_template(self):
        """Download Excel template for Optical products"""
        self.ensure_one()
        try:
            template_data = self._get_import_template_data('opt')
            filename = f"BangNhap_Gong_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return self._create_download_action(template_data, filename)
        except Exception as e:
            _logger.error(f"Error generating optical template: {e}", exc_info=True)
            raise UserError(_('Error generating template: %s') % str(e))
    
    def action_download_accessory_template(self):
        """Download Excel template for Accessory products"""
        self.ensure_one()
        try:
            template_data = self._get_import_template_data('accessory')
            filename = f"BangNhap_PhuKien_{datetime.now().strftime('%Y%m%d')}.xlsx"
            return self._create_download_action(template_data, filename)
        except Exception as e:
            _logger.error(f"Error generating accessory template: {e}", exc_info=True)
            raise UserError(_('Error generating template: %s') % str(e))

    def _get_import_template_data(self, product_type):
        if product_type == 'lens':
            raw_data = excel_template_generator.generate_lens_template()
        elif product_type == 'opt':
            raw_data = excel_template_generator.generate_opt_template()
        elif product_type == 'accessory':
            raw_data = excel_template_generator.generate_accessory_template()
        else:
            raise UserError(_('Unsupported template type: %s') % product_type)
        return self._ensure_template_supplier_currency_columns(raw_data)

    def _ensure_template_supplier_currency_columns(self, template_data):
        """Ensure download template always contains supplier_ref and currency_id columns."""
        try:
            from openpyxl import load_workbook
        except Exception:
            return template_data

        wb = load_workbook(filename=BytesIO(template_data))
        ws = wb.active

        tech_row = None
        for row_idx in range(1, min(ws.max_row, 20) + 1):
            values = [
                str(cell.value).strip()
                for cell in ws[row_idx]
                if cell.value not in (None, '')
            ]
            if {'FullName', 'TradeMark', 'Group'}.intersection(values):
                tech_row = row_idx
                break

        if not tech_row:
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        label_row = tech_row - 1 if tech_row > 1 else tech_row
        existing_keys = {
            str(cell.value).strip().lower()
            for cell in ws[tech_row]
            if cell.value not in (None, '')
        }

        columns_to_add = []
        if not any(k in existing_keys for k in ('supplier', 'supplier_ref', 'partner_ref')):
            columns_to_add.append(('Nhà cung cấp (ref)', 'supplier_ref'))
        if not any(k in existing_keys for k in ('currency', 'currency_id')):
            columns_to_add.append(('Đơn vị nguyên tệ', 'currency_id'))

        next_col = ws.max_column + 1
        for label, key in columns_to_add:
            ws.cell(row=label_row, column=next_col, value=label)
            ws.cell(row=tech_row, column=next_col, value=key)
            next_col += 1

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _create_download_action(self, file_data, filename):
        """Create download action for Excel file"""
        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(file_data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
    
    
    def action_parse_excel(self):
        """
        Parse uploaded Excel file and move to preview state
        """
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file first.'))
        
        try:
            # Decode file
            file_content = base64.b64decode(self.excel_file)
            
            # Parse Excel
            parsed_data = excel_reader.parse_excel_file(file_content, self.file_name or 'uploaded.xlsx')

            self._normalize_import_rows(parsed_data.get('rows') or [])
            
            # Set product type
            self.product_type = parsed_data['product_type']

            cache = data_cache.MasterDataCache(self.env)

            # Ensure rows can be imported without manual default_code column.
            self._populate_missing_default_codes(
                parsed_data['rows'],
                parsed_data['product_type'],
                cache=cache,
                strict=False,
            )
            
            # Validate data
            validation_result = import_validator.validate_all_rows(
                self.env,
                cache,
                parsed_data['rows'],
                parsed_data['product_type']
            )
            
            # Clear existing preview lines
            self.preview_line_ids.unlink()
            
            # Create preview lines with auto-generated codes
            preview_lines = self._generate_preview_lines(
                parsed_data['rows'],
                parsed_data['product_type'],
                cache,
                validation_result
            )
            self.preview_line_ids = preview_lines
            
            # Keep JSON for debugging
            preview = {
                'product_type': parsed_data['product_type'],
                'total_rows': parsed_data['total_rows'],
                'headers': parsed_data['headers'],
                'validation': {
                    'valid': validation_result['valid'],
                    'error_count': len(validation_result['errors']),
                    'warning_count': len(validation_result['warnings']),
                }
            }
            self.preview_data = json.dumps(preview, indent=2, ensure_ascii=False)
            
            # Format error log
            if validation_result['errors'] or validation_result['warnings']:
                error_lines = []
                
                if validation_result['errors']:
                    error_lines.append('ERRORS:')
                    error_lines.append('=' * 60)
                    for err in validation_result['errors']:
                        row_info = f"Row {err['row']}: " if err['row'] else ""
                        error_lines.append(f"  {row_info}{err['message']}")
                
                if validation_result['warnings']:
                    error_lines.append('\nWARNINGS:')
                    error_lines.append('=' * 60)
                    for warn in validation_result['warnings']:
                        row_info = f"Row {warn['row']}: " if warn['row'] else ""
                        error_lines.append(f"  {row_info}{warn['message']}")
                
                self.error_log = '\n'.join(error_lines)
            else:
                self.error_log = False
            
            # Store full data in context for import (temporary solution)
            # In production, might want to use ir.attachment or database storage
            self = self.with_context(parsed_excel_data=parsed_data)
            
            # Move to preview state
            self.state = 'preview'
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'product.excel.import',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'context': {'parsed_excel_data': parsed_data}
            }
        
        except Exception as e:
            _logger.error(f"Error parsing Excel file: {str(e)}", exc_info=True)
            raise UserError(_('Error parsing Excel file: %s') % str(e))
    
    def action_back_to_upload(self):
        self.ensure_one()
        self.state = 'upload'
        self.error_log = False
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'product.excel.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_confirm_import(self):
        """
        Import products into database using batch processing for optimal performance.
        
        Optimizations:
        - Batch create: Creates all products in batches instead of one-by-one
        - Disabled tracking: Turns off mail/activity tracking during import
        - Pre-generated codes: Generates all product codes before creating products
        - Batch commits: Commits every BATCH_SIZE records to avoid long locks
        """
        self.ensure_one()
        
        BATCH_SIZE = 100  # Commit every 100 records
        
        # Get parsed data from context (temporary solution)
        parsed_data = self.env.context.get('parsed_excel_data')
        if not parsed_data:
            # Try to re-parse if data lost
            _logger.warning("Parsed data not found in context, re-parsing...")
            file_content = base64.b64decode(self.excel_file)
            parsed_data = excel_reader.parse_excel_file(file_content, self.file_name or 'uploaded.xlsx')

        self._normalize_import_rows(parsed_data.get('rows') or [])
        
        cache = data_cache.MasterDataCache(self.env)

        # Ensure missing default_code values are auto-generated before validation/import.
        # Strict mode provides explicit errors for missing Group/Brand/Index in import rows.
        self._populate_missing_default_codes(
            parsed_data['rows'],
            parsed_data['product_type'],
            cache=cache,
            strict=True,
        )

        # Validate again
        validation_result = import_validator.validate_all_rows(
            self.env,
            cache,
            parsed_data['rows'],
            parsed_data['product_type']
        )
        
        if not validation_result['valid']:
            raise UserError(_(
                'Cannot import: Data validation failed. Please fix errors and try again.\n\n'
                'See Error Log tab for details.'
            ))
        
        # Apply optimization context flags
        optimized_self = self.with_context(
            tracking_disable=True,
            mail_notrack=True,
            mail_create_nolog=True,
            no_reset_password=True,
            import_mode=True,
            prefetch_fields=False,
        )
        
        rows = parsed_data['rows']
        product_type = parsed_data['product_type']
        total_rows = len(rows)
        
        _logger.info(f"Starting batch import of {total_rows} {product_type} products...")
        
        success_count = 0
        error_count = 0
        error_messages = []
        
        try:
            # Process in batches
            for batch_start in range(0, total_rows, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, total_rows)
                batch_rows = rows[batch_start:batch_end]
                batch_num = (batch_start // BATCH_SIZE) + 1
                total_batches = (total_rows + BATCH_SIZE - 1) // BATCH_SIZE
                
                _logger.info(f"Processing batch {batch_num}/{total_batches} (rows {batch_start + 1}-{batch_end})")
                
                try:
                    with optimized_self.env.cr.savepoint():
                        created_count = optimized_self._create_products_batch(
                            batch_rows, product_type, cache
                        )
                        success_count += created_count
                    
                    # Commit after successful batch
                    optimized_self.env.cr.commit()
                    _logger.info(f"Batch {batch_num} committed: {created_count} products created")
                    
                except Exception as e:
                    error_count += len(batch_rows)
                    error_messages.append(f"Batch {batch_num} failed: {str(e)}")
                    _logger.error(f"Error in batch {batch_num}: {str(e)}", exc_info=True)
                    # Continue with next batch
            
            if error_count > 0 and success_count == 0:
                raise UserError(_(
                    'Import failed completely. All batches had errors.\n\n%s'
                ) % '\n'.join(error_messages[:10]))
        
        except Exception as e:
            self.error_log = str(e)
            self.error_count = error_count
            raise
        
        # Success (possibly partial)
        self.success_count = success_count
        self.error_count = error_count
        
        if error_count > 0:
            self.error_log = f"Imported {success_count} products with {error_count} errors:\n" + '\n'.join(error_messages[:10])
        else:
            self.error_log = f"Successfully imported {success_count} products!"
        
        self.state = 'done'
        
        _logger.info(f"Import completed: {success_count} success, {error_count} errors")
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'product.excel.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _get_row_value(self, row_data, keys):
        if not isinstance(row_data, dict):
            return False
        lower_map = {str(k).strip().lower(): k for k in row_data.keys()}
        for key in keys:
            source_key = key if key in row_data else lower_map.get(str(key).strip().lower())
            if source_key is None:
                continue
            value = row_data.get(source_key)
            if value in (None, False):
                continue
            if isinstance(value, str):
                value = value.strip()
            if value == '':
                continue
            return value
        return False

    def _normalize_row_token(self, value):
        if value in (None, False):
            return ''
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_import_rows(self, rows):
        for row in rows:
            if not isinstance(row, dict):
                continue
            if not row.get('Supplier'):
                supplier_alias = self._get_row_value(row, ['supplier_ref', 'SupplierRef', 'supplier', 'partner_ref'])
                if supplier_alias:
                    row['Supplier'] = self._normalize_row_token(supplier_alias)
            if not row.get('Currency'):
                currency_alias = self._get_row_value(row, ['currency_id', 'CurrencyId', 'currency', 'Loại tiền tệ', 'Đơn vị nguyên tệ'])
                if currency_alias:
                    row['Currency'] = self._normalize_row_token(currency_alias)
            if not row.get('TradeMark'):
                brand_alias = self._get_row_value(row, ['brand_id', 'Brand', 'brand'])
                if brand_alias:
                    row['TradeMark'] = self._normalize_row_token(brand_alias)
            if not row.get('Group'):
                group_alias = self._get_row_value(row, ['group_id', 'group'])
                if group_alias:
                    row['Group'] = self._normalize_row_token(group_alias)
            if not row.get('Index'):
                index_alias = self._get_row_value(row, ['lens_index_id', 'index_id', 'index'])
                if index_alias:
                    row['Index'] = self._normalize_row_token(index_alias)

    def _resolve_currency_record(self, row_data, cache, row_number=None, strict=False):
        token = self._get_row_value(row_data, ['Currency', 'currency_id', 'currency', 'Loại tiền tệ', 'Đơn vị nguyên tệ'])
        token = self._normalize_row_token(token)
        if not token:
            return False

        currency = cache.get_currency(token) if hasattr(cache, 'get_currency') else False
        if currency:
            return currency

        Currency = self.env['res.currency'].with_context(active_test=False)
        search_domain = ['|', ('name', '=', token), ('name', '=', token.upper())]
        records = Currency.search(search_domain)
        # Also try full_name and symbol if available
        if not records and 'full_name' in Currency._fields:
            records = Currency.search(['|', ('full_name', '=', token), ('full_name', '=', token.upper())])
        if not records and 'symbol' in Currency._fields:
            records = Currency.search([('symbol', '=', token)])

        if not records or len(records) == 0:
            if strict:
                row_label = f"Dòng {row_number}" if row_number else "Dòng import"
                raise ValidationError(f"{row_label}: Không tìm thấy tiền tệ với giá trị '{token}'.")
            return False
        if len(records) > 1:
            if strict:
                row_label = f"Dòng {row_number}" if row_number else "Dòng import"
                raise ValidationError(f"{row_label}: Có nhiều loại tiền tệ cùng mã/code '{token}'. Vui lòng kiểm tra lại dữ liệu tiền tệ.")
            return False
        return records[0]

    def _resolve_supplier_partner(self, row_data, cache, row_number=None, strict=False):
        token = self._get_row_value(row_data, ['Supplier', 'supplier_ref', 'SupplierRef', 'supplier', 'partner_ref'])
        token = self._normalize_row_token(token)
        if not token:
            return False

        Partner = self.env['res.partner'].with_context(active_test=False)
        # Only match supplier with exact ref and supplier_rank > 0
        domain = [('ref', '=', token), ('supplier_rank', '>', 0)]
        records = Partner.search(domain)
        if not records or len(records) == 0:
            if strict:
                row_label = f"Dòng {row_number}" if row_number else "Dòng import"
                raise ValidationError(f"{row_label}: Không tìm thấy nhà cung cấp theo ref '{token}'.")
            return False
        if len(records) > 1:
            if strict:
                row_label = f"Dòng {row_number}" if row_number else "Dòng import"
                raise ValidationError(f"{row_label}: Có nhiều nhà cung cấp cùng ref '{token}'. Vui lòng kiểm tra lại danh sách nhà cung cấp.")
            return False
        return records[0]

    def _resolve_code_generation_inputs(self, row_data, product_type, cache, row_number=None, strict=False):
        row_label = f"Dòng {row_number}" if row_number else "Dòng import"

        group_token = self._get_row_value(row_data, ['Group', 'group_id', 'group'])
        group_token = self._normalize_row_token(group_token)
        group = cache.get_group(group_token) if group_token else False
        if strict and not group:
            raise ValidationError(f"{row_label}: Thiếu hoặc không tìm thấy Group để sinh mã sản phẩm.")

        brand_token = self._get_row_value(row_data, ['TradeMark', 'brand_id', 'Brand', 'brand'])
        brand_token = self._normalize_row_token(brand_token)
        brand = cache.get_brand(brand_token) if brand_token else False
        if strict and not brand:
            raise ValidationError(f"{row_label}: Thiếu hoặc không tìm thấy TradeMark để sinh mã sản phẩm.")

        lens_index_id = False
        if product_type == 'lens':
            index_token = self._get_row_value(row_data, ['Index', 'lens_index_id', 'index_id', 'index'])
            index_token = self._normalize_row_token(index_token)
            index_record = cache.get_lens_index(index_token) if index_token else False
            if strict and not index_record:
                raise ValidationError(f"{row_label}: Lens thiếu hoặc không tìm thấy Index để sinh mã sản phẩm.")
            if index_record:
                lens_index_id = index_record.id

        return group.id if group else False, brand.id if brand else False, lens_index_id

    def _generate_default_code_for_row(self, row_data, product_type, cache, sequence_cache, row_number=None, strict=False):
        group_id, brand_id, lens_index_id = self._resolve_code_generation_inputs(
            row_data,
            product_type,
            cache,
            row_number=row_number,
            strict=strict,
        )

        if not group_id or not brand_id:
            return False
        if product_type == 'lens' and not lens_index_id:
            return False

        code = self.env['product.template'].generate_next_default_code(
            group_id=group_id,
            brand_id=brand_id,
            lens_index_id=lens_index_id,
            product_type='lens' if product_type == 'lens' else 'non_lens',
            sequence_cache=sequence_cache,
        )
        return code

    def _populate_missing_default_codes(self, rows, product_type, cache, strict=False):
        sequence_cache = {}
        generated_codes = []
        for idx, row_data in enumerate(rows, start=1):
            existing_code = (row_data.get('default_code') or '').strip().upper()
            if existing_code:
                generated_codes.append(existing_code)
                row_data['default_code'] = existing_code
                continue

            code = self._generate_default_code_for_row(
                row_data,
                product_type,
                cache,
                sequence_cache=sequence_cache,
                row_number=idx,
                strict=strict,
            )
            generated_codes.append(code)
            if code and not row_data.get('default_code'):
                row_data['default_code'] = code
        return generated_codes
    
    def _generate_preview_lines(self, rows, product_type, cache, validation_result):
        """Generate preview lines from parsed Excel data"""
        
        preview_lines = []
        error_dict = {}
        
        # Build error dictionary by row number
        for err in validation_result.get('errors', []):
            row = err.get('row')
            if row:
                if row not in error_dict:
                    error_dict[row] = []
                error_dict[row].append(err['message'])

        sequence_cache = {}
        
        for idx, row_data in enumerate(rows, start=1):
            generated_code = ''
            try:
                generated_code = (row_data.get('default_code') or '').strip().upper()
                if not generated_code:
                    generated_code = self._generate_default_code_for_row(
                        row_data,
                        product_type,
                        cache,
                        sequence_cache=sequence_cache,
                        row_number=idx,
                        strict=False,
                    ) or 'Thiếu dữ liệu sinh mã'
            except Exception as e:
                _logger.warning(f"Could not generate code for row {idx}: {e}")
                generated_code = f"ERROR: {str(e)[:50]}"
            
            # Build preview line
            line_vals = {
                'row_number': idx,
                'full_name': row_data.get('FullName', ''),
                'eng_name': row_data.get('EngName', ''),
                'group': row_data.get('Group', ''),
                'brand': row_data.get('TradeMark', ''),
                'generated_code': generated_code,
                'retail_price': float(row_data.get('Retail_Price', 0) or 0),
                'wholesale_price': float(row_data.get('Wholesale_Price', 0) or 0),
                'cost_price': float(row_data.get('Cost_Price', 0) or 0),
            }
            
            # Add type-specific fields
            if product_type == 'lens':
                line_vals.update({
                    'index': row_data.get('Index', ''),
                    'design1': row_data.get('Design1', ''),
                    'material': row_data.get('Material', ''),
                })
            elif product_type == 'opt':
                line_vals.update({
                    'sku': row_data.get('Sku', ''),
                    'model': row_data.get('Model', ''),
                    'frame_type': row_data.get('Frame_Type', ''),
                })
            
            # Add errors if any
            excel_row = row_data.get('_excel_row', idx)
            if excel_row in error_dict:
                line_vals['has_error'] = True
                line_vals['error_message'] = '\n'.join(error_dict[excel_row])
            
            preview_lines.append((0, 0, line_vals))
        
        return preview_lines
    
    def _create_products_batch(self, rows, product_type, cache):
        """
        Create multiple products at once using batch operations.
        
        This is much faster than creating products one by one because:
        - Uses batch code generation (1 query per prefix instead of per product)
        - Uses Odoo's batch create (1 insert per batch instead of per product)
        - Pre-processes all data before any database operations
        
        Args:
            rows: List of row data dictionaries
            product_type: 'lens', 'opt', or 'accessory'
            cache: MasterDataCache instance
            
        Returns:
            Number of successfully created products
        """
        if not rows:
            return 0

        # Phase 1: Generate all codes at once (with per-prefix batch cache)
        generated_codes = self._populate_missing_default_codes(
            rows,
            product_type,
            cache,
            strict=False,
        )
        
        # Phase 2: Prepare all product vals
        all_product_vals = []
        for idx, row_data in enumerate(rows):
            product_vals = self._prepare_product_vals(row_data, product_type, cache)
            
            # Apply generated code
            code = generated_codes[idx]
            if not code:
                raise ValidationError(f"Dòng {idx + 1}: Không thể sinh default_code cho sản phẩm import.")
            product_vals['default_code'] = code
            product_vals['auto_generate_code'] = False
            
            all_product_vals.append(product_vals)
        
        # Phase 3: Batch create products
        products = self.env['product.template'].create(all_product_vals)
        
        _logger.debug(f"Batch created {len(products)} products")
        
        return len(products)
    
    def _prepare_product_vals(self, row_data, product_type, cache):
        """
        Prepare product values dictionary from row data.
        Extracted from _create_product for reuse in batch operations.
        """
        product_vals = {
            'name': row_data.get('FullName'),
            'eng_name': row_data.get('EngName'),
            'type': 'consu',
        }

        group_id, brand_id, _lens_index_id = self._resolve_code_generation_inputs(
            row_data,
            product_type,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )
        product_vals['group_id'] = group_id
        product_vals['brand_id'] = brand_id
        
        # Optional foreign keys
        # Supplier - use seller_ids (Odoo standard)
        currency = self._resolve_currency_record(
            row_data,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )

        supplier = self._resolve_supplier_partner(
            row_data,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )
        if supplier:
            seller_vals = {
                'partner_id': supplier.id,
                'price': float(row_data.get('Origin_Price', 0) or 0),
                'min_qty': 1.0,
                'delay': 1,
            }
            if currency:
                seller_vals['currency_id'] = currency.id
            product_vals['seller_ids'] = [(0, 0, seller_vals)]
        
        if row_data.get('Country'):
            country = cache.get_country(row_data['Country'])
            if country:
                product_vals['country_id'] = country.id
        
        if row_data.get('Warranty'):
            warranty = cache.get_warranty(row_data['Warranty'])
            if warranty:
                product_vals['warranty_id'] = warranty.id
        
        if currency:
            if 'currency_zone_id' in self.env['product.template']._fields:
                product_vals['currency_zone_id'] = currency.id
            elif 'currency_id' in self.env['product.template']._fields:
                product_vals['currency_id'] = currency.id
        
        # Prices
        product_vals['or_price'] = float(row_data.get('Origin_Price', 0) or 0)
        product_vals['standard_price'] = float(row_data.get('Cost_Price', 0) or 0)
        product_vals['list_price'] = float(row_data.get('Retail_Price', 0) or 0)
        product_vals['ws_price'] = float(row_data.get('Wholesale_Price', 0) or 0)
        product_vals['ws_price_max'] = float(row_data.get('Wholesale_Price_Max', 0) or 0)
        product_vals['ws_price_min'] = float(row_data.get('Wholesale_Price_Min', 0) or 0)
        
        # Text fields
        if row_data.get('Unit'):
            product_vals['unit'] = row_data['Unit']
        if row_data.get('Use'):
            product_vals['uses'] = row_data['Use']
        if row_data.get('Guide'):
            product_vals['guide'] = row_data['Guide']
        if row_data.get('Warning'):
            product_vals['warning'] = row_data['Warning']
        if row_data.get('Preserve'):
            product_vals['preserve'] = row_data['Preserve']
        if row_data.get('Description'):
            product_vals['description'] = row_data['Description']
        if row_data.get('Note'):
            product_vals['description_sale'] = row_data['Note']
        
        # Employee assignment
        if self.employee_id:
            product_vals['employee_id'] = self.employee_id.id
        
        # Image
        if row_data.get('Image'):
            product_vals['image_1920'] = row_data['Image']
        
        # Create lens/opt specific data
        if product_type == 'lens':
            product_vals['lens_ids'] = [(0, 0, self._prepare_lens_vals(row_data, cache))]
        elif product_type == 'opt':
            product_vals['opt_ids'] = [(0, 0, self._prepare_opt_vals(row_data, cache))]
        
        return product_vals
    
    def _create_product(self, row_data, product_type, cache):

        product_vals = {
            'name': row_data.get('FullName'),
            'eng_name': row_data.get('EngName'),
            'type': 'product',  # Stockable product
        }

        # Group (required)
        group_id, brand_id, _lens_index_id = self._resolve_code_generation_inputs(
            row_data,
            product_type,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )
        product_vals['group_id'] = group_id

        # Brand (required)
        product_vals['brand_id'] = brand_id

        # Optional foreign keys
        currency = self._resolve_currency_record(
            row_data,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )

        supplier = self._resolve_supplier_partner(
            row_data,
            cache,
            row_number=row_data.get('_excel_row'),
            strict=True,
        )
        if supplier:
            seller_vals = {
                'partner_id': supplier.id,
                'price': float(row_data.get('Origin_Price', 0) or 0),
                'min_qty': 1.0,
                'delay': 1,
            }
            if currency:
                seller_vals['currency_id'] = currency.id
            product_vals['seller_ids'] = [(0, 0, seller_vals)]

        if row_data.get('Country'):
            country = cache.get_country(row_data['Country'])
            if country:
                product_vals['country_id'] = country.id

        if row_data.get('Warranty'):
            warranty = cache.get_warranty(row_data['Warranty'])
            if warranty:
                product_vals['warranty_id'] = warranty.id

        if currency:
            if 'currency_zone_id' in self.env['product.template']._fields:
                product_vals['currency_zone_id'] = currency.id
            elif 'currency_id' in self.env['product.template']._fields:
                product_vals['currency_id'] = currency.id

        # Prices
        product_vals['or_price'] = float(row_data.get('Origin_Price', 0) or 0)
        product_vals['standard_price'] = float(row_data.get('Cost_Price', 0) or 0)
        product_vals['list_price'] = float(row_data.get('Retail_Price', 0) or 0)
        product_vals['ws_price'] = float(row_data.get('Wholesale_Price', 0) or 0)
        product_vals['ws_price_max'] = float(row_data.get('Wholesale_Price_Max', 0) or 0)
        product_vals['ws_price_min'] = float(row_data.get('Wholesale_Price_Min', 0) or 0)
        
        # Text fields
        if row_data.get('Unit'):
            product_vals['unit'] = row_data['Unit']
        if row_data.get('Use'):
            product_vals['uses'] = row_data['Use']
        if row_data.get('Guide'):
            product_vals['guide'] = row_data['Guide']
        if row_data.get('Warning'):
            product_vals['warning'] = row_data['Warning']
        if row_data.get('Preserve'):
            product_vals['preserve'] = row_data['Preserve']
        if row_data.get('Description'):
            product_vals['description'] = row_data['Description']
        if row_data.get('Note'):
            product_vals['description_sale'] = row_data['Note']
        
        # Image
        if row_data.get('Image'):
            product_vals['image_1920'] = row_data['Image']
        
        # Generate default_code for single-row flow
        code = self._generate_default_code_for_row(
            row_data,
            product_type,
            cache,
            sequence_cache={},
            row_number=row_data.get('_excel_row'),
            strict=True,
        )
        product_vals['default_code'] = code
        product_vals['auto_generate_code'] = False
        
        # Create lens/opt specific data
        if product_type == 'lens':
            product_vals['lens_ids'] = [(0, 0, self._prepare_lens_vals(row_data, cache))]
        elif product_type == 'opt':
            product_vals['opt_ids'] = [(0, 0, self._prepare_opt_vals(row_data, cache))]
        
        # Create product
        product = self.env['product.template'].create(product_vals)
        
        return product
    
    def _prepare_lens_vals(self, row_data, cache):
        """Prepare lens_ids values"""
        lens_vals = {}
        
        # Simple text fields
        for excel_field, odoo_field in [
            ('SPH', 'sph'), ('CYL', 'cyl'), ('ADD', 'len_add'),
            ('AXIS', 'axis'), ('PRISM', 'prism'), ('PRISMBASE', 'prism_base'),
            ('BASE', 'base'), ('Abbe', 'abbe'), ('Polarized', 'polarized'),
            ('Diameter', 'diameter'), ('ColorInt', 'color_int'),
            ('Corridor', 'corridor'), ('MirCoating', 'mir_coating'),
        ]:
            if row_data.get(excel_field):
                lens_vals[odoo_field] = row_data[excel_field]
        
        # Foreign keys
        if row_data.get('Design1'):
            design = cache.get_design(row_data['Design1'])
            if design:
                lens_vals['design1_id'] = design.id
        
        if row_data.get('Design2'):
            design = cache.get_design(row_data['Design2'])
            if design:
                lens_vals['design2_id'] = design.id
        
        if row_data.get('Material'):
            material = cache.get_material(row_data['Material'])
            if material:
                lens_vals['material_id'] = material.id
        
        if row_data.get('Index'):
            index = cache.get_lens_index(row_data['Index'])
            if index:
                lens_vals['index_id'] = index.id
        
        if row_data.get('Uv'):
            uv = cache.get_uv(row_data['Uv'])
            if uv:
                lens_vals['uv_id'] = uv.id
        
        if row_data.get('HMC'):
            hmc = cache.get_color(row_data['HMC'])
            if hmc:
                lens_vals['cl_hmc_id'] = hmc.id
        
        if row_data.get('PHO'):
            pho = cache.get_color(row_data['PHO'])
            if pho:
                lens_vals['cl_pho_id'] = pho.id
        
        if row_data.get('TIND'):
            tind = cache.get_color(row_data['TIND'])
            if tind:
                lens_vals['cl_tint_id'] = tind.id
        
        # Many2many: Coating (CSV)
        if row_data.get('Coating'):
            coating_ids = []
            for coating_cid in str(row_data['Coating']).split(','):
                coating = cache.get_coating(coating_cid.strip())
                if coating:
                    coating_ids.append(coating.id)
            if coating_ids:
                lens_vals['coating_ids'] = [(6, 0, coating_ids)]
        
        return lens_vals
    
    def _prepare_opt_vals(self, row_data, cache):
        """Prepare opt_ids values"""
        opt_vals = {}
        
        # Simple text fields
        for excel_field, odoo_field in [
            ('Sku', 'sku'), ('Model', 'model'), ('Model_Supplier', 'oem_ncc'),
            ('Serial', 'serial'), ('Season', 'season'),
        ]:
            if row_data.get(excel_field):
                opt_vals[odoo_field] = row_data[excel_field]
        
        # Gender
        if row_data.get('Gender'):
            opt_vals['gender'] = str(row_data['Gender'])
        
        # Dimensions
        for excel_field, odoo_field in [
            ('Lens_Width', 'lens_width'), ('Bridge_Width', 'bridge_width'),
            ('Temple_Width', 'temple_width'), ('Lens_Height', 'lens_height'),
            ('Lens_Span', 'lens_span'),
        ]:
            if row_data.get(excel_field):
                try:
                    opt_vals[odoo_field] = int(row_data[excel_field])
                except (ValueError, TypeError):
                    pass
        
        # Foreign keys
        fk_mapping = [
            ('Frame', 'frame_id', cache.get_frame),
            ('Frame_Type', 'frame_type_id', cache.get_frame_type),
            ('Shape', 'shape_id', cache.get_shape),
            ('Ve', 've_id', cache.get_ve),
            ('Temple', 'temple_id', cache.get_temple),
            ('Material_Ve', 'material_ve_id', cache.get_material),
            ('Material_TempleTip', 'material_temple_tip_id', cache.get_material),
            ('Material_Lens', 'material_lens_id', cache.get_material),
            ('Color_Lens', 'color_lens_id', cache.get_color),
            ('Color_Opt_Front', 'color_front_id', cache.get_color),
            ('Color_Opt_Temple', 'color_temple_id', cache.get_color),
        ]
        
        for excel_field, odoo_field, getter in fk_mapping:
            if row_data.get(excel_field):
                record = getter(row_data[excel_field])
                if record:
                    opt_vals[odoo_field] = record.id
        
        # Many2many: Materials (CSV)
        if row_data.get('Material_Opt_Front'):
            material_ids = []
            for mat_cid in str(row_data['Material_Opt_Front']).split(','):
                material = cache.get_material(mat_cid.strip())
                if material:
                    material_ids.append(material.id)
            if material_ids:
                opt_vals['materials_front_ids'] = [(6, 0, material_ids)]
        
        if row_data.get('Material_Opt_Temple'):
            material_ids = []
            for mat_cid in str(row_data['Material_Opt_Temple']).split(','):
                material = cache.get_material(mat_cid.strip())
                if material:
                    material_ids.append(material.id)
            if material_ids:
                opt_vals['materials_temple_ids'] = [(6, 0, material_ids)]
        
        # Many2many: Coating (CSV)
        if row_data.get('Coating'):
            coating_ids = []
            for coating_cid in str(row_data['Coating']).split(','):
                coating = cache.get_coating(coating_cid.strip())
                if coating:
                    coating_ids.append(coating.id)
            if coating_ids:
                opt_vals['coating_ids'] = [(6, 0, coating_ids)]
        
        return opt_vals
