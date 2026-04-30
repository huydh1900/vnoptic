# -*- coding: utf-8 -*-
import base64
import io
import logging
import re

from openpyxl import load_workbook

from odoo import api, fields, models, _
from odoo.exceptions import UserError

from ..models.product_import_base_inherit import (
    VNOP_TEMPLATE_FIELD_CODE_OVERRIDES,
    VNOP_TEMPLATE_VIRTUAL_IMPORT_COLUMNS,
)

_logger = logging.getLogger(__name__)

PRODUCT_TYPE_SELECTION = [
    ('auto', 'Theo Mã nhóm hàng'),
]


# ─────────────────────────────────────────────────────────────────────────
#   MAPPING NHÃN TIẾNG VIỆT  →  TECHNICAL FIELD NAME
#   Áp dụng cho format Excel "ItemLst - quy ước PM" (header nhãn tiếng Việt).
#   Nếu cùng một field có nhiều biến thể tiếng Việt, liệt kê ở dạng tuple.
# ─────────────────────────────────────────────────────────────────────────
VN_LABEL_TO_FIELD = {
    # Common
    'mã hàng tự định nghĩa': 'default_code',
    'mã hàng tự đinh nghĩa': 'default_code',
    'mã nhóm hàng': 'classification_code',
    'xuất xứ': 'country_code',
    'ma ncc': 'supplier_ref',
    'mã ncc': 'supplier_ref',
    'thương hiệu': 'brand_code',
    'tên đầy đủ': 'name',
    'tên rút gọn': 'x_short_name',
    'tên viết hóa đơn': 'x_invoice_name',
    'mã vạch': 'barcode',
    'mã đơn vị tối thiểu': 'uom_name',
    'giá nhập chỉ định chưa thuế': 'standard_price',
    'giá bán cơ sở có thuế': 'x_base_price',
    'giá bán lẻ có thuế': 'list_price',
    'giá bán buôn có thuế': 'x_ws_price',
    'thuế suất mua vào': 'supplier_taxes_code',
    'mã thuế bán ra': 'taxes_code',
    'kiểu in nhãn': 'label_print_type',
    'phụ kiện': 'accessory_note',
    # Frame
    'mã model': 'opt_model',
    'mã màu': 'opt_color',
    'loại gọng': 'opt_frame_type_cid',
    'cấu trúc vành': 'opt_frame_structure_cid',
    'giới tính': 'opt_gender',
    'dáng mắt': 'opt_shape_cid',
    'loại ve': 'opt_ve_cid',
    'loại chuôi càng': 'opt_temple_tip_cid',
    'chất liệu mặt trước': 'opt_materials_front_cid',
    'chất liệu càng': 'opt_materials_temple_cid',
    'trọng lượng (với gọng kính bằng vàng)': 'opt_weight',
    'trọng lượng': 'opt_weight',
    'số lượng đá quý (kim cương/saphire) - kèm carat': 'opt_gem_combined',
    'số lượng đá quý': 'opt_gem_combined',
    'chất liệu ve': 'opt_material_ve_cid',
    'chất liệu chuôi càng': 'opt_material_temple_tip_cid',
    'chất liệu mắt': 'opt_material_lens_cid',
    'mạ mắt polarized': 'opt_polarized',
    'lớp tráng mắt': 'opt_coating_cid',
    'màu sắc mặt trước': 'opt_color_front',
    'màu sắc càng kính': 'opt_color_temple',
    'màu sắc mắt kính': 'opt_color_lens',
    'ngang mắt': 'ngang_mat',
    'cao mắt': 'cao_mat',
    'dài cầu': 'opt_bridge_width',
    'dài càng': 'opt_temple_width',
    # Lens
    'loai tròng': 'len_type',
    'loại tròng': 'len_type',
    'cứng/mềm với loại áp tròng': 'lens_hard_soft',
    'đặc tính': 'lens_features',
    'độ cầu': 'x_sph',
    'độ trụ': 'x_cyl',
    'độ cộng thêm': 'x_add',
    'trái/phải': 'lens_eye_side',
    'trục': 'x_axis',
    'lăng kính': 'x_prism',
    'đáy lăng kính': 'x_prism_base',
    'độ cong kính': 'lens_base_curve',
    'chỉ số tán sắc': 'lens_abbe',
    'hạng mục tròng kính': 'lens_category',
    'đường kính': 'x_diameter',
    'thiết kế': 'lens_design_cid',
    'lớp film chức năng': 'lens_film_cid',
    'chất liệu': 'lens_material_code',
    'chiết suất': 'lens_index_name',
    'chống uv': 'lens_uv_cid',
    'tỷ lệ ánh sáng truyền qua': 'lens_light_transmission',
    'lớp phủ': 'lens_coating_cid',
    'ánh mạ': 'lens_mirror_coating',
    'đổi màu': 'lens_cl_pho_cid',
    'mạ màu': 'lens_color_coating',
    'độ đậm màu': 'lens_color_int',
    'corridor': 'lens_corridor',
    'phủ gương': 'lens_mirror_color',
}

# Cột bị ghi chú "bỏ đi" hoặc không cần map
VN_LABEL_IGNORE = {
    'mã số hàng tự sinh',           # auto-generated
    'loại hàng',                     # luôn 01 - hình thức quản lý cũ
    'tên nguồn cung cấp',           # tự suy ra từ Ma NCC
    'hình thức quản lý (01)',       # legacy
    'phân cực',                      # bỏ đi (note row 1)
    'hiêu suất phân cực',
    'hiệu suất phân cực',
    'trục phân cực',
}

# Map giá trị nhãn → selection code (cho field selection)
LEN_TYPE_VALUE_MAP = {
    'sv': 'SV', 'bf': 'BF', 'tf': 'TF', 'pro': 'PRO', 'pt': 'PT',
    'dt': 'SV', 'ht': 'BF', 'dat': 'PRO',  # legacy alias
    'đơn tròng': 'SV', 'hai tròng': 'BF', 'ba tròng': 'TF',
    'đa tròng': 'PRO', 'phôi tròng': 'PT',
}

GENDER_VALUE_MAP = {
    'm': 'M', 'f': 'F', 'u': 'U',
    'k-m': 'K-M', 'k-f': 'K-F', 'k-u': 'K-U',
    'nam': 'M', 'nữ': 'F', 'unisex': 'U',
    'trẻ em - nam': 'K-M', 'trẻ em - nữ': 'K-F', 'trẻ em - unisex': 'K-U',
    '1': 'M', '2': 'F', '3': 'U',
}

EYE_SIDE_VALUE_MAP = {
    'trái': 'left', 'phải': 'right', 'cả hai': 'both',
    'l': 'left', 'r': 'right', 'lr': 'both',
    'left': 'left', 'right': 'right', 'both': 'both',
}

HARD_SOFT_VALUE_MAP = {
    'cứng': 'hard', 'mềm': 'soft',
    'hard': 'hard', 'soft': 'soft',
}

LABEL_PRINT_VALUE_MAP = {
    '0': '0', '1': '1', '2': '2', '3': '3',
    '0 - không in': '0',
    '1- mẫu in tròng kính': '1',
    '2- mẫu in gọng kính': '2',
    'không in': '0',
    'mẫu in tròng kính': '1',
    'mẫu in gọng kính': '2',
}


def _norm_label(text):
    """Chuẩn hoá nhãn tiếng Việt: lowercase, strip space thừa."""
    if not text:
        return ''
    text = str(text).strip().lower()
    text = re.sub(r'\s+', ' ', text)
    return text


# ─────────────────────────────────────────────────────────────────────────
#   Map field-key (column token) → (model, search_field, label_field, model_label)
#   Dùng chung cho validation + quick-create + lookup.
# ─────────────────────────────────────────────────────────────────────────
REF_SPECS = {
    'country_code': ('res.country', 'code', 'Xuất xứ', 'Quốc gia'),
    'brand_code': ('product.brand', 'code', 'Thương hiệu', 'Thương hiệu'),
    'uom_name': ('uom.uom', 'name', 'Đơn vị tính', 'Đơn vị đo'),
    'classification_code': ('product.classification', 'code', 'Mã nhóm hàng', 'Nhóm sản phẩm'),
    'lens_index_name': ('product.lens.index', 'name', 'Chiết suất', 'Chiết suất'),
    'lens_design_cid': ('product.design', 'code', 'Thiết kế tròng', 'Thiết kế tròng'),
    'lens_material_code': ('product.lens.material', 'code', 'Chất liệu tròng', 'Chất liệu tròng'),
    'lens_film_cid': ('product.lens.film', 'cid', 'Lớp film', 'Lớp film chức năng'),
    'lens_uv_cid': ('product.uv', 'cid', 'Chống UV', 'Chống UV'),
    'lens_coating_cid': ('product.coating', 'cid', 'Lớp phủ tròng', 'Lớp phủ'),
    'lens_cl_pho_cid': ('product.lens.photochromic', 'cid', 'Đổi màu', 'Đổi màu'),
    'opt_frame_type_cid': ('product.frame.type', 'cid', 'Loại gọng', 'Loại gọng'),
    'opt_frame_structure_cid': ('product.frame.structure', 'cid', 'Cấu trúc vành', 'Cấu trúc vành'),
    'opt_shape_cid': ('product.shape', 'cid', 'Dáng mắt', 'Dáng mắt'),
    'opt_ve_cid': ('product.ve', 'cid', 'Loại ve', 'Loại ve'),
    'opt_temple_tip_cid': ('product.temple.tip', 'cid', 'Loại chuôi càng', 'Loại chuôi càng'),
    'opt_material_ve_cid': ('product.material', 'cid', 'Chất liệu ve', 'Chất liệu'),
    'opt_material_lens_cid': ('product.material', 'cid', 'Chất liệu mắt', 'Chất liệu'),
    'opt_material_temple_tip_cid': ('product.material', 'cid', 'Chất liệu chuôi càng', 'Chất liệu'),
    'opt_materials_front_cid': ('product.material', 'cid', 'Chất liệu mặt trước', 'Chất liệu'),
    'opt_materials_temple_cid': ('product.material', 'cid', 'Chất liệu càng', 'Chất liệu'),
    'opt_coating_cid': ('product.coating', 'cid', 'Lớp tráng mắt', 'Lớp phủ'),
}

# Các field key có nhiều giá trị (csv) trong cell
REF_MULTI_KEYS = {
    'lens_design_cid', 'lens_material_code', 'lens_film_cid', 'lens_coating_cid',
    'opt_material_temple_tip_cid', 'opt_materials_front_cid', 'opt_materials_temple_cid',
    'opt_coating_cid',
}


class ProductImportMissingRef(models.TransientModel):
    """Mỗi dòng = 1 giá trị chưa có trong master data, kèm nút Tạo nhanh."""
    _name = 'product.import.missing.ref'
    _description = 'Master data thiếu khi import sản phẩm'

    wizard_id = fields.Many2one('product.import.wizard', ondelete='cascade', required=True)
    field_key = fields.Char(string='Cột Excel', readonly=True,
                            help='Token cột trong file Excel (vd: brand_code).')
    field_label = fields.Char(string='Trường', readonly=True)
    model_name = fields.Char(string='Model', readonly=True)
    model_label = fields.Char(string='Bảng dữ liệu', readonly=True)
    search_field = fields.Char(string='Khóa tra cứu', readonly=True,
                               help='Field dùng để tra cứu trên model (cid/code/name).')
    value = fields.Char(string='Giá trị', required=True,
                        help='Giá trị xuất hiện trong Excel mà chưa có trên hệ thống.')
    display_name_value = fields.Char(string='Tên hiển thị',
                                     help='Tên đầy đủ sẽ ghi vào field "name" của bảng. Mặc định = giá trị Excel.')
    state = fields.Selection([
        ('pending', 'Chưa tạo'),
        ('created', 'Đã tạo'),
    ], default='pending', readonly=True)
    created_record_display = fields.Char(string='Bản ghi đã tạo', readonly=True)

    @api.depends('value')
    def _onchange_default_display_name(self):
        # khi nhập value lần đầu, copy sang display_name_value
        pass

    def action_quick_create(self):
        """Tạo bản ghi master data tương ứng."""
        for rec in self:
            if rec.state == 'created':
                continue
            if not rec.value:
                raise UserError(_('Giá trị không được trống.'))
            if rec.model_name not in self.env:
                raise UserError(_('Model %s không tồn tại.') % rec.model_name)
            CoModel = self.env[rec.model_name].sudo()

            # Build vals: ghi search_field=value và name=display_name_value (nếu model có name)
            vals = {}
            if rec.search_field and rec.search_field in CoModel._fields:
                vals[rec.search_field] = rec.value
            if 'name' in CoModel._fields:
                vals['name'] = rec.display_name_value or rec.value
            elif rec.search_field != 'name' and 'name' not in vals:
                # Trường hợp model không có 'name', chỉ ghi search_field
                pass

            # Tránh duplicate: nếu đã tồn tại record có search_field=value thì link luôn
            existing = False
            if rec.search_field and rec.search_field in CoModel._fields:
                existing = CoModel.search([(rec.search_field, '=', rec.value)], limit=1)
            if existing:
                rec.write({
                    'state': 'created',
                    'created_record_display': '%s [đã có sẵn]' % existing.display_name,
                })
                continue

            try:
                new_rec = CoModel.create(vals)
            except Exception as e:
                raise UserError(_('Không tạo được %s với giá trị "%s": %s')
                                % (rec.model_label or rec.model_name, rec.value, str(e)))
            rec.write({
                'state': 'created',
                'created_record_display': new_rec.display_name,
            })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'product.import.wizard',
            'res_id': self.wizard_id.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_open_record_form(self):
        """Mở form sửa bản ghi đầy đủ (cho trường hợp cần thêm thông tin khác)."""
        self.ensure_one()
        if self.state != 'created':
            raise UserError(_('Bản ghi chưa được tạo.'))
        if self.model_name not in self.env:
            raise UserError(_('Model %s không tồn tại.') % self.model_name)
        CoModel = self.env[self.model_name]
        rec = CoModel.search([(self.search_field, '=', self.value)], limit=1)
        if not rec:
            raise UserError(_('Không tìm thấy bản ghi.'))
        return {
            'type': 'ir.actions.act_window',
            'res_model': self.model_name,
            'res_id': rec.id,
            'view_mode': 'form',
            'target': 'new',
        }


class ProductImportWizard(models.TransientModel):
    _name = 'product.import.wizard'
    _description = 'Import sản phẩm từ Excel'

    product_type = fields.Selection(
        selection=PRODUCT_TYPE_SELECTION,
        string='Loại sản phẩm',
        required=True,
        default='auto',
        help='"Tự động" suy loại sản phẩm theo cột "Mã nhóm hàng" trong file (đề xuất cho file ItemLst).'
    )
    file_data = fields.Binary(string='File Excel', required=True)
    file_name = fields.Char()

    state = fields.Selection([
        ('upload', 'Upload'),
        ('preview', 'Kiểm thử'),
        ('running', 'Đang import'),
        ('done', 'Hoàn tất'),
    ], default='upload')
    # Progress tracking (cập nhật theo batch trong background job)
    progress_total = fields.Integer(string='Tổng số dòng', readonly=True)
    progress_done = fields.Integer(string='Đã xử lý', readonly=True)
    progress_percent = fields.Float(
        string='Tiến độ (%)', compute='_compute_progress_percent',
        readonly=True, store=False,
    )
    progress_message = fields.Char(string='Trạng thái xử lý', readonly=True)
    imported_count = fields.Integer(string='Số sản phẩm đã import', readonly=True)
    imported_lens_count = fields.Integer(string='Tròng', readonly=True)
    imported_frame_count = fields.Integer(string='Gọng', readonly=True)
    imported_accessory_count = fields.Integer(string='Phụ kiện', readonly=True)
    imported_other_count = fields.Integer(string='Khác', readonly=True)
    imported_product_ids = fields.Many2many(
        'product.template',
        string='Sản phẩm đã tạo',
        readonly=True,
    )
    error_text = fields.Text(string='Chi tiết lỗi', readonly=True)
    preview_text = fields.Html(string='Kết quả kiểm thử', readonly=True, sanitize=False)
    missing_ref_ids = fields.One2many(
        'product.import.missing.ref', 'wizard_id',
        string='Master data thiếu',
    )
    missing_ref_count = fields.Integer(
        string='SL master thiếu', compute='_compute_missing_ref_count',
    )

    @api.depends('missing_ref_ids', 'missing_ref_ids.state')
    def _compute_missing_ref_count(self):
        for r in self:
            r.missing_ref_count = sum(
                1 for m in r.missing_ref_ids if m.state == 'pending'
            )

    @api.depends('progress_total', 'progress_done')
    def _compute_progress_percent(self):
        for r in self:
            r.progress_percent = (r.progress_done / r.progress_total * 100.0) if r.progress_total else 0.0

    # Ngưỡng dòng tối đa import đồng bộ (request HTTP). Vượt ngưỡng → đẩy queue_job.
    _IMPORT_SYNC_THRESHOLD = 200
    # Kích thước batch khi tạo product.template hàng loạt.
    _IMPORT_BATCH_SIZE = 200

    def action_refresh_progress(self):
        """Reload wizard form (dùng cho nút Cập nhật ở state 'running')."""
        self.ensure_one()
        return self._reopen()

    _MAX_SAMPLE = 5

    _FIELD_LOOKUP_CONFIG = {
        'categ_id': ('code', []),
        'brand_id': ('code', []),
        'country_id': ('code', []),
        'uom_id': ('name', []),
        'warranty_id': ('code', []),
        'warranty_supplier_id': ('code', []),
        'warranty_retail_id': ('code', []),
        'lens_design_ids': ('cid', []),
        'lens_material_ids': ('code', []),
        'lens_film_ids': ('cid', []),
        'lens_index_id': ('name', []),
        'lens_uv_id': ('cid', []),
        'lens_cl_hmc_id': ('cid', []),
        'lens_cl_pho_id': ('cid', []),
        'lens_cl_tint_id': ('cid', []),
        'opt_frame_type_id': ('cid', []),
        'opt_frame_structure_id': ('cid', []),
        'opt_shape_id': ('cid', []),
        'opt_ve_id': ('cid', []),
        'opt_temple_tip_id': ('cid', []),
        'opt_material_ve_id': ('cid', []),
        'opt_material_temple_tip_ids': ('cid', []),
        'opt_material_lens_id': ('cid', []),
        'opt_materials_front_ids': ('cid', []),
        'opt_materials_temple_ids': ('cid', []),
        'opt_coating_ids': ('cid', []),
        'lens_coating_ids': ('cid', []),
        'design_id': ('cid', []),
        'shape_id': ('cid', []),
        'material_id': ('cid', []),
        'color_id': ('cid', []),
    }

    _CODE_TO_FIELD = {v: k for k, v in VNOP_TEMPLATE_FIELD_CODE_OVERRIDES.items()}

    # ────────────────────────────────────────────────────────────
    #   ENTRY ACTIONS
    # ────────────────────────────────────────────────────────────

    def action_test(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))

        raw = base64.b64decode(self.file_data)
        fmt = self._detect_format(raw)
        if fmt == 'vn_label':
            return self._action_test_vn_label(raw)
        return self._action_test_technical(raw)

    def action_import(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Vui lòng chọn file Excel.'))

        raw = base64.b64decode(self.file_data)
        fmt = self._detect_format(raw)
        if fmt == 'vn_label':
            return self._action_import_vn_label(raw)
        return self._action_import_technical(raw)

    def action_back_upload(self):
        self.ensure_one()
        self.missing_ref_ids.unlink()
        self.write({
            'state': 'upload',
            'preview_text': False,
            'imported_count': 0,
            'imported_lens_count': 0,
            'imported_frame_count': 0,
            'imported_accessory_count': 0,
            'imported_other_count': 0,
            'imported_product_ids': [(5, 0, 0)],
            'error_text': False,
            'progress_total': 0,
            'progress_done': 0,
            'progress_message': False,
        })
        return self._reopen()

    def action_create_all_missing(self):
        """Tạo nhanh hàng loạt tất cả master data còn thiếu."""
        self.ensure_one()
        pending = self.missing_ref_ids.filtered(lambda r: r.state == 'pending')
        if not pending:
            raise UserError(_('Không còn master data nào cần tạo.'))
        pending.action_quick_create()
        # Sau khi tạo xong, chạy lại Kiểm thử để cập nhật cảnh báo còn lại
        return self.action_test()

    def action_view_imported_products(self):
        """Mở list view sản phẩm đã tạo (lọc theo classification_type)."""
        self.ensure_one()
        if not self.imported_product_ids:
            raise UserError(_('Chưa có sản phẩm nào được import.'))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Sản phẩm vừa import'),
            'res_model': 'product.template',
            'view_mode': 'list,form',
            'domain': [('id', 'in', self.imported_product_ids.ids)],
            'context': {'group_by': ['classification_type']},
        }

    # ────────────────────────────────────────────────────────────
    #   FORMAT DETECTION
    # ────────────────────────────────────────────────────────────

    def _detect_format(self, file_bytes):
        """'vn_label' nếu có dòng chứa nhãn TV (≥3 nhãn), ngược lại 'technical'."""
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            return 'technical'
        ws = wb.active
        if ws is None:
            wb.close()
            return 'technical'

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 20:
                break
            cells = [_norm_label(c) for c in row if c not in (None, '')]
            matches = sum(1 for c in cells if c in VN_LABEL_TO_FIELD)
            if matches >= 5:
                wb.close()
                return 'vn_label'
        wb.close()
        return 'technical'

    # ────────────────────────────────────────────────────────────
    #   VN-LABEL FORMAT: PARSE
    # ────────────────────────────────────────────────────────────

    def _parse_vn_label_excel(self, file_bytes):
        """Parse Excel format VN-label → (header_fields, data_rows, errors)."""
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            return [], [], [_('Không đọc được file Excel: %s') % str(e)]

        ws = wb.active
        if ws is None:
            wb.close()
            return [], [], [_('File Excel không có sheet nào.')]

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([self._normalize_cell(c) for c in row])
        wb.close()

        if not all_rows:
            return [], [], [_('File Excel trống.')]

        header_index = None
        best_score = 0
        for idx, row in enumerate(all_rows):
            if idx > 20:
                break
            cells = [_norm_label(c) for c in row]
            score = sum(1 for c in cells if c in VN_LABEL_TO_FIELD)
            if score > best_score and score >= 5:
                best_score = score
                header_index = idx

        if header_index is None:
            return [], [], [_('Không tìm thấy dòng header nhãn tiếng Việt (vd: "Tên đầy đủ", "Mã nhóm hàng").')]

        raw_header = all_rows[header_index]
        # Map từng cột: nhãn → field name | None (skip)
        header_fields = []
        for cell in raw_header:
            norm = _norm_label(cell)
            if not norm:
                header_fields.append(None)
            elif norm in VN_LABEL_IGNORE:
                header_fields.append(None)
            elif norm in VN_LABEL_TO_FIELD:
                header_fields.append(VN_LABEL_TO_FIELD[norm])
            else:
                header_fields.append(None)  # cột không nhận diện → bỏ qua

        # Data start: dòng tiếp theo có "Tên đầy đủ" hoặc "Mã nhóm hàng" non-empty
        try:
            name_col = header_fields.index('name')
        except ValueError:
            name_col = None
        try:
            classif_col = header_fields.index('classification_code')
        except ValueError:
            classif_col = None

        data_rows = []
        for row in all_rows[header_index + 1:]:
            cells = list(row) + [''] * max(0, len(header_fields) - len(row))
            cells = cells[:len(header_fields)]
            # Bỏ dòng trống
            if all(not self._normalize_cell(c) for c in cells):
                continue
            # Bỏ note rows (chứa cụm từ note đặc trưng)
            joined = ' '.join(self._normalize_cell(c).lower() for c in cells)
            if any(marker in joined for marker in [
                'cho phép tự điền', 'được phép ngăn cách', 'my thêm cột', 'ghi màu phủ gương',
            ]) and not (name_col is not None and self._normalize_cell(cells[name_col])):
                continue
            # Phải có ít nhất name HOẶC classification để là data row
            has_name = name_col is not None and self._normalize_cell(cells[name_col])
            has_classif = classif_col is not None and self._normalize_cell(cells[classif_col])
            if not has_name and not has_classif:
                continue
            data_rows.append([self._normalize_cell(c) for c in cells])

        return header_fields, data_rows, []

    # ────────────────────────────────────────────────────────────
    #   VN-LABEL FORMAT: TEST (preview)
    # ────────────────────────────────────────────────────────────

    def _action_test_vn_label(self, raw):
        header, rows, errors = self._parse_vn_label_excel(raw)
        if errors:
            self.write({
                'state': 'preview',
                'preview_text': self._html_error_panel(_('Lỗi đọc file'), '<br/>'.join(errors)),
            })
            return self._reopen()
        if not rows:
            self.write({
                'state': 'preview',
                'preview_text': self._html_error_panel(
                    _('Lỗi'), _('File không có dữ liệu hợp lệ.')),
            })
            return self._reopen()

        issues, classification_buckets = self._validate_vn_rows(header, rows)
        html = self._build_vn_preview_html(header, rows, issues, classification_buckets)
        self.write({'state': 'preview', 'preview_text': html})
        return self._reopen()

    def _validate_vn_rows(self, header, rows):
        """Trả về (issues, buckets) — buckets={lens: [n], frame: [n], accessory: [n], other: [n]}."""
        issues = []
        buckets = {'lens': [], 'frame': [], 'accessory': [], 'other': []}

        col_index = {f: i for i, f in enumerate(header) if f}

        # 1. name bắt buộc
        if 'name' not in col_index:
            issues.append({'level': 'error', 'title': _('Thiếu cột "Tên đầy đủ" trong file')})
            return issues, buckets
        name_idx = col_index['name']

        # 2. Suy classification_type cho từng dòng
        Classification = self.env['product.classification']
        classif_cache = {
            (c.code or '').strip(): c for c in Classification.search([])
        }

        empty_name_rows = []
        unknown_classif = {}
        duplicate_names = {}
        seen_names = {}

        for r_idx, row in enumerate(rows):
            row_no = r_idx + 1
            name = (row[name_idx] or '').strip()
            if not name:
                empty_name_rows.append(row_no)
                continue

            # Duplicate trong file
            if name in seen_names:
                duplicate_names.setdefault(name, [seen_names[name]]).append(row_no)
            else:
                seen_names[name] = row_no

            # Suy category_type
            ctype = 'other'
            if 'classification_code' in col_index:
                code = (row[col_index['classification_code']] or '').strip()
                if code:
                    code_norm = code.lstrip('0') or '0'
                    classif = classif_cache.get(code) or classif_cache.get(code_norm)
                    if classif:
                        ctype = classif.category_type or 'other'
                    else:
                        unknown_classif[code] = unknown_classif.get(code, 0) + 1
            buckets[ctype].append(row_no)

        if empty_name_rows:
            sample = empty_name_rows[:self._MAX_SAMPLE]
            extra = len(empty_name_rows) - self._MAX_SAMPLE
            detail = _('Dòng: %s') % ', '.join(str(r) for r in sample)
            if extra > 0:
                detail += _(' ... và %s dòng khác') % extra
            issues.append({
                'level': 'error',
                'title': _('Tên đầy đủ trống ở %s dòng') % len(empty_name_rows),
                'details': [detail],
            })

        if duplicate_names:
            details = []
            sample = list(duplicate_names.keys())[:self._MAX_SAMPLE]
            for nm in sample:
                rows_str = ', '.join(str(r) for r in duplicate_names[nm][:5])
                details.append('"%s" (dòng: %s)' % (nm, rows_str))
            issues.append({
                'level': 'error',
                'title': _('Tên trùng trong file: %s tên') % len(duplicate_names),
                'details': details,
            })

        if unknown_classif:
            details = ['"%s" (×%s)' % (k, v) for k, v in list(unknown_classif.items())[:self._MAX_SAMPLE]]
            issues.append({
                'level': 'warn',
                'title': _('Mã nhóm hàng không tìm thấy: %s mã') % len(unknown_classif),
                'details': details,
            })

        # 3. Kiểm tra name đã tồn tại trong DB
        unique_names = list(seen_names.keys())
        if unique_names:
            existing = self.env['product.template'].with_context(active_test=False).search_read(
                [('name', 'in', unique_names)], ['name']
            )
            existing_names = [r['name'] for r in existing if r['name']]
            if existing_names:
                details = [', '.join(existing_names[:10])]
                if len(existing_names) > 10:
                    details.append(_('... và %s tên khác') % (len(existing_names) - 10))
                issues.append({
                    'level': 'warn',
                    'title': _('Tên đã tồn tại trong hệ thống: %s tên (sẽ bỏ qua khi import)')
                             % len(existing_names),
                    'details': details,
                })

        # 4. Kiểm tra giá trị M2O/M2M tồn tại
        self._validate_vn_relational(issues, header, rows, col_index)

        return issues, buckets

    def _validate_vn_relational(self, issues, header, rows, col_index):
        """Phát hiện giá trị cid/code không tồn tại → đẩy vào missing_ref_ids."""
        # Xóa danh sách cũ để tránh nhân đôi khi user ấn Kiểm thử nhiều lần
        self.missing_ref_ids.unlink()

        existing_pairs = set()  # (field_key, value) — tránh trùng giữa các lần
        new_missing_lines = []
        total_missing = 0

        for token, (model, search_field, field_label, model_label) in REF_SPECS.items():
            if token not in col_index:
                continue
            idx = col_index[token]
            multi = token in REF_MULTI_KEYS
            values = set()
            for row in rows:
                raw = (row[idx] or '').strip()
                if not raw:
                    continue
                if multi:
                    for v in re.split(r'[,;]', raw):
                        v = v.strip()
                        if v:
                            values.add(v)
                else:
                    values.add(raw)

            if not values:
                continue
            if model not in self.env:
                continue
            CoModel = self.env[model]
            if search_field not in CoModel._fields:
                continue
            # Áp alias trước khi tra cứu (vd: CAI → Cái)
            resolved_values = set()
            for v in values:
                if token == 'uom_name' and v in self._UOM_ALIASES:
                    resolved_values.add(self._UOM_ALIASES[v])
                else:
                    resolved_values.add(v)
            found = CoModel.search_read(
                [(search_field, 'in', list(resolved_values))], [search_field])
            found_set = {r[search_field] for r in found if r.get(search_field)}
            # Brand: cho phép match thêm theo name (CID hoặc Tên đều hợp lệ)
            extra_name_set = set()
            if model == 'product.brand':
                by_name = CoModel.search_read(
                    [('name', 'in', list(resolved_values))], ['name'])
                extra_name_set = {r['name'] for r in by_name if r.get('name')}
            # Đánh dấu missing dựa trên giá trị gốc; nếu giá trị gốc đã có alias hợp lệ → coi như tìm thấy
            missing = []
            for v in sorted(values):
                resolved = self._UOM_ALIASES.get(v, v) if token == 'uom_name' else v
                if resolved not in found_set and resolved not in extra_name_set:
                    missing.append(v)
            if not missing:
                continue
            total_missing += len(missing)
            for value in missing:
                pair = (token, value)
                if pair in existing_pairs:
                    continue
                existing_pairs.add(pair)
                new_missing_lines.append((0, 0, {
                    'field_key': token,
                    'field_label': field_label,
                    'model_name': model,
                    'model_label': model_label,
                    'search_field': search_field,
                    'value': value,
                    'display_name_value': value,
                }))

        if new_missing_lines:
            self.write({'missing_ref_ids': new_missing_lines})

        if total_missing:
            issues.append({
                'level': 'warn',
                'title': _('Có %s giá trị master data chưa tồn tại') % total_missing,
                'details': [_(
                    'Xem chi tiết và bấm "Tạo nhanh" ngay bên dưới để bổ sung trước khi import.'
                )],
            })

    # ────────────────────────────────────────────────────────────
    #   VN-LABEL FORMAT: IMPORT
    # ────────────────────────────────────────────────────────────

    def _action_import_vn_label(self, raw):
        # Parse + validate nhanh ở foreground để cho user feedback ngay nếu file lỗi.
        header, rows, errors = self._parse_vn_label_excel(raw)
        if errors:
            self.write({
                'state': 'done', 'imported_count': 0,
                'error_text': '\n'.join(errors),
            })
            return self._reopen()
        if not rows:
            self.write({
                'state': 'done', 'imported_count': 0,
                'error_text': _('File không có dữ liệu hợp lệ.'),
            })
            return self._reopen()
        col_index = {f: i for i, f in enumerate(header) if f}
        if 'name' not in col_index:
            self.write({
                'state': 'done', 'imported_count': 0,
                'error_text': _('Thiếu cột "Tên đầy đủ" trong file.'),
            })
            return self._reopen()

        # Chặn import nếu file có trùng tên/barcode hoặc đã tồn tại trong DB.
        dup_errors = self._assert_no_duplicates(rows, col_index)
        if dup_errors:
            self.write({
                'state': 'done', 'imported_count': 0,
                'error_text': _('Không thể import do trùng tên/mã vạch:') + '\n\n' + '\n\n'.join(dup_errors),
            })
            return self._reopen()

        total = len(rows)
        # File nhỏ → chạy đồng bộ (vẫn batched để giảm overhead) để user nhận kết quả ngay.
        if total <= self._IMPORT_SYNC_THRESHOLD:
            self.write({
                'state': 'running', 'progress_total': total, 'progress_done': 0,
                'progress_message': _('Đang import...'),
                'imported_count': 0, 'error_text': False,
            })
            self._run_vn_label_import(header, rows, col_index)
            return self._reopen()

        # File lớn → enqueue queue_job, UI chuyển sang state 'running' với progressbar.
        self.write({
            'state': 'running', 'progress_total': total, 'progress_done': 0,
            'progress_message': _('Đã đưa vào hàng đợi, đang chờ xử lý...'),
            'imported_count': 0,
            'imported_lens_count': 0, 'imported_frame_count': 0,
            'imported_accessory_count': 0, 'imported_other_count': 0,
            'imported_product_ids': [(5, 0, 0)],
            'error_text': False,
        })
        # Commit để background worker đọc được state mới + để vượt qua giới hạn HTTP.
        self.env.cr.commit()
        raw_b64 = base64.b64encode(raw).decode('ascii')
        self.with_delay(
            description=_('Import VN-label: %s (%s dòng)') % (self.file_name or '?', total),
        )._import_vn_label_job(raw_b64)
        return self._reopen()

    def _run_vn_label_import(self, header, rows, col_index):
        """Logic import thực tế. Cập nhật progress sau mỗi batch và commit để
        UI (state, progress_done) phản ánh tiến độ ngay khi user reload form.
        Dùng chung cho cả flow đồng bộ và queue_job worker.
        """
        self.ensure_one()
        caches = self._build_lookup_caches()

        all_names = list({(r[col_index['name']] or '').strip() for r in rows
                          if (r[col_index['name']] or '').strip()})
        existing_names = set()
        if all_names:
            existing_names = set(
                self.env['product.template'].with_context(active_test=False).search([
                    ('name', 'in', all_names)
                ]).mapped('name')
            )

        # Barcode duy nhất: pre-fetch các barcode đã tồn tại trong DB.
        bc_idx = col_index.get('barcode')
        dc_idx = col_index.get('default_code')  # VN label dùng default_code → barcode
        all_barcodes = set()
        for r in rows:
            for ix in (bc_idx, dc_idx):
                if ix is None:
                    continue
                v = (r[ix] or '').strip() if ix < len(r) else ''
                if v:
                    all_barcodes.add(v)
        existing_barcodes = set()
        if all_barcodes:
            existing_barcodes = set(
                self.env['product.template'].with_context(active_test=False).search([
                    ('barcode', 'in', list(all_barcodes))
                ]).mapped('barcode')
            )

        # Tắt mail tracking + chatter để giảm chi phí khi tạo hàng loạt.
        Product = self.env['product.template'].with_context(
            tracking_disable=True,
            mail_create_nolog=True,
            mail_notrack=True,
        )

        created_ids = []
        skipped_existing = 0
        skipped_barcode = 0
        errors_per_row = []
        batch_vals = []
        batch_meta = []  # [(row_idx, name)]
        batch_size = self._IMPORT_BATCH_SIZE
        total = len(rows)

        def _flush(batch_vals, batch_meta):
            if not batch_vals:
                return
            try:
                recs = Product.create(batch_vals)
                created_ids.extend(recs.ids)
            except Exception:
                # Fallback row-by-row để xác định dòng lỗi cụ thể, không phá hỏng cả batch.
                for v, (ri, nm) in zip(batch_vals, batch_meta):
                    try:
                        with self.env.cr.savepoint():
                            rec = Product.create(v)
                        created_ids.append(rec.id)
                    except Exception as ee:
                        errors_per_row.append('Dòng %s ("%s"): %s' % (ri + 1, nm, ee))

        for r_idx, row in enumerate(rows):
            try:
                vals = self._row_to_vals(row, col_index, caches)
            except Exception as e:
                errors_per_row.append('Dòng %s: %s' % (r_idx + 1, str(e)))
                continue
            if not vals.get('name'):
                continue
            if vals['name'] in existing_names:
                skipped_existing += 1
                continue
            bc = (vals.get('barcode') or '').strip()
            if bc and bc in existing_barcodes:
                skipped_barcode += 1
                errors_per_row.append(
                    'Dòng %s ("%s"): mã vạch "%s" đã tồn tại, bỏ qua.' % (
                        r_idx + 1, vals.get('name'), bc,
                    )
                )
                continue
            existing_names.add(vals['name'])  # tránh trùng trong cùng file
            if bc:
                existing_barcodes.add(bc)  # tránh trùng barcode trong cùng file
            batch_vals.append(vals)
            batch_meta.append((r_idx, vals['name']))

            if len(batch_vals) >= batch_size:
                _flush(batch_vals, batch_meta)
                batch_vals, batch_meta = [], []
                done = r_idx + 1
                self.write({
                    'progress_done': done,
                    'progress_message': _('Đã xử lý %s/%s dòng (%s sản phẩm)') % (
                        done, total, len(created_ids),
                    ),
                })
                # Commit theo batch: persist progress + giải phóng locks.
                self.env.cr.commit()

        # Flush phần còn lại.
        _flush(batch_vals, batch_meta)

        # Counters phân loại
        created = self.env['product.template'].browse(created_ids) if created_ids \
            else self.env['product.template']
        lens_n = sum(1 for p in created if p.classification_type == 'lens')
        frame_n = sum(1 for p in created if p.classification_type == 'frame')
        acc_n = sum(1 for p in created if p.classification_type == 'accessory')
        other_n = len(created) - lens_n - frame_n - acc_n

        error_text_parts = []
        if skipped_existing:
            error_text_parts.append(_('Đã bỏ qua %s sản phẩm trùng tên.') % skipped_existing)
        if skipped_barcode:
            error_text_parts.append(_('Đã bỏ qua %s sản phẩm trùng mã vạch.') % skipped_barcode)
        if errors_per_row:
            error_text_parts.append(_('Lỗi từng dòng:') + '\n' + '\n'.join(errors_per_row[:50]))
            if len(errors_per_row) > 50:
                error_text_parts.append(_('... và %s lỗi khác') % (len(errors_per_row) - 50))

        self.write({
            'state': 'done',
            'progress_done': total,
            'progress_message': _('Hoàn tất.'),
            'imported_count': len(created_ids),
            'imported_lens_count': lens_n,
            'imported_frame_count': frame_n,
            'imported_accessory_count': acc_n,
            'imported_other_count': other_n,
            'imported_product_ids': [(6, 0, created_ids)],
            'error_text': '\n\n'.join(error_text_parts) if error_text_parts else False,
        })

    def _import_vn_label_job(self, raw_b64):
        """Worker queue_job: chạy import VN-label trong background."""
        self.ensure_one()
        try:
            raw = base64.b64decode(raw_b64)
            header, rows, errors = self._parse_vn_label_excel(raw)
            if errors or not rows:
                self.write({
                    'state': 'done',
                    'imported_count': 0,
                    'error_text': '\n'.join(errors) if errors else _('File không có dữ liệu hợp lệ.'),
                })
                return
            col_index = {f: i for i, f in enumerate(header) if f}
            if 'name' not in col_index:
                self.write({
                    'state': 'done', 'imported_count': 0,
                    'error_text': _('Thiếu cột "Tên đầy đủ" trong file.'),
                })
                return
            # Re-validate trùng vì DB có thể đã đổi giữa lúc enqueue và lúc chạy.
            dup_errors = self._assert_no_duplicates(rows, col_index)
            if dup_errors:
                self.write({
                    'state': 'done', 'imported_count': 0,
                    'error_text': _('Không thể import do trùng tên/mã vạch:') + '\n\n' + '\n\n'.join(dup_errors),
                })
                return
            self._run_vn_label_import(header, rows, col_index)
        except Exception as e:
            _logger.exception('Import VN-label job lỗi')
            self.write({
                'state': 'done',
                'progress_message': _('Lỗi.'),
                'error_text': str(e),
            })

    # ────────────────────────────────────────────────────────────
    #   VN-LABEL: ROW → VALS
    # ────────────────────────────────────────────────────────────

    # Alias: giá trị Excel ngắn / không dấu → tên chính thức trong master data
    _UOM_ALIASES = {
        'cai': 'Cái', 'cái': 'Cái', 'CAI': 'Cái', 'CÁI': 'Cái',
        'pcs': 'Cái', 'piece': 'Cái', 'pc': 'Cái',
    }

    def _build_lookup_caches(self):
        env = self.env

        def cache_by(model, key):
            if model not in env:
                return {}
            CoModel = env[model]
            if key not in CoModel._fields:
                return {}
            res = {}
            for r in CoModel.search([]):
                v = r[key]
                if v:
                    res[str(v).strip()] = r.id
            return res

        # UoM cache + alias (CAI / cai → Cái)
        uom_cache = cache_by('uom.uom', 'name')
        for alias, canonical in self._UOM_ALIASES.items():
            if canonical in uom_cache and alias not in uom_cache:
                uom_cache[alias] = uom_cache[canonical]

        # Brand: cho phép tra cứu theo cả code (CID) lẫn name
        brand_cache = {}
        for r in env['product.brand'].search([]):
            if r.code:
                brand_cache[str(r.code).strip()] = r.id
            if r.name:
                brand_cache.setdefault(str(r.name).strip(), r.id)

        return {
            'classification_code': cache_by('product.classification', 'code'),
            'category_code': cache_by('product.category', 'code'),
            'country_code': cache_by('res.country', 'code'),
            'brand_code': brand_cache,
            'uom_name': uom_cache,
            'lens_index_name': cache_by('product.lens.index', 'name'),
            'lens_design_cid': cache_by('product.design', 'code'),
            'lens_material_code': cache_by('product.lens.material', 'code'),
            'lens_film_cid': cache_by('product.lens.film', 'cid'),
            'lens_uv_cid': cache_by('product.uv', 'cid'),
            'lens_coating_cid': cache_by('product.coating', 'cid'),
            'lens_cl_pho_cid': cache_by('product.lens.photochromic', 'cid'),
            'opt_frame_type_cid': cache_by('product.frame.type', 'cid'),
            'opt_frame_structure_cid': cache_by('product.frame.structure', 'cid'),
            'opt_shape_cid': cache_by('product.shape', 'cid'),
            'opt_ve_cid': cache_by('product.ve', 'cid'),
            'opt_temple_tip_cid': cache_by('product.temple.tip', 'cid'),
            'opt_material_cid': cache_by('product.material', 'cid'),
            'supplier_ref': {
                p.ref.strip(): p.id for p in env['res.partner'].search([
                    ('supplier_rank', '>', 0), ('ref', '!=', False),
                ]) if p.ref
            },
            'taxes_name': {
                t.name.strip(): t.id for t in env['account.tax'].search([
                    ('type_tax_use', '=', 'sale'),
                ]) if t.name
            },
            'supplier_taxes_name': {
                t.name.strip(): t.id for t in env['account.tax'].search([
                    ('type_tax_use', '=', 'purchase'),
                ]) if t.name
            },
        }

    def _row_to_vals(self, row, col_index, caches):
        """Convert một dòng Excel thành dict vals cho product.template.create()."""
        vals = {}

        def cell(token):
            idx = col_index.get(token)
            if idx is None:
                return ''
            return (row[idx] or '').strip()

        def cell_csv(token):
            v = cell(token)
            if not v:
                return []
            return [x.strip() for x in re.split(r'[,;]', v) if x.strip()]

        # Common scalar
        # Cột Excel "default_code" → ghi vào field barcode (CID là khóa nội bộ)
        if cell('default_code'):
            vals['barcode'] = cell('default_code')
        if cell('name'):
            vals['name'] = cell('name')
        if cell('x_short_name'):
            vals['x_short_name'] = cell('x_short_name')
        if cell('x_invoice_name'):
            vals['x_invoice_name'] = cell('x_invoice_name')
        if cell('barcode'):
            vals['barcode'] = cell('barcode')
        if cell('accessory_note'):
            vals['accessory_note'] = cell('accessory_note')

        # Numeric
        for tok, field_name in [
            ('standard_price', 'standard_price'),
            ('list_price', 'list_price'),
            ('x_base_price', 'x_base_price'),
            ('x_ws_price', 'x_ws_price'),
            ('opt_weight', 'opt_weight'),
            ('lens_base_curve', 'lens_base_curve'),
            ('lens_abbe', 'lens_abbe'),
            ('lens_corridor', 'lens_corridor'),
            ('lens_light_transmission', 'lens_light_transmission'),
            ('x_diameter', 'x_diameter'),
            ('x_add', 'x_add'),
            ('x_prism', 'x_prism'),
            ('ngang_mat', 'ngang_mat'),
            ('cao_mat', 'cao_mat'),
        ]:
            v = cell(tok)
            if v:
                try:
                    vals[field_name] = float(str(v).replace(',', '.'))
                except (TypeError, ValueError):
                    pass

        for tok, field_name in [
            ('opt_bridge_width', 'opt_bridge_width'),
            ('opt_temple_width', 'opt_temple_width'),
            ('x_axis', 'x_axis'),
        ]:
            v = cell(tok)
            if v:
                try:
                    vals[field_name] = int(float(str(v).replace(',', '.')))
                except (TypeError, ValueError):
                    pass

        # Frame char fields
        for tok in [
            'opt_model', 'opt_color', 'opt_color_front', 'opt_color_temple', 'opt_color_lens',
            'lens_features', 'lens_category', 'lens_color_int', 'lens_mirror_coating',
            'lens_color_coating', 'lens_mirror_color', 'x_prism_base',
        ]:
            v = cell(tok)
            if v:
                vals[tok] = v

        # Selection: Loại tròng / giới tính / trái phải / cứng mềm / kiểu in nhãn
        v = cell('len_type').lower()
        if v in LEN_TYPE_VALUE_MAP:
            vals['len_type'] = LEN_TYPE_VALUE_MAP[v]
        v = cell('opt_gender').lower()
        if v in GENDER_VALUE_MAP:
            vals['opt_gender'] = GENDER_VALUE_MAP[v]
        v = cell('lens_eye_side').lower()
        if v in EYE_SIDE_VALUE_MAP:
            vals['lens_eye_side'] = EYE_SIDE_VALUE_MAP[v]
        v = cell('lens_hard_soft').lower()
        if v in HARD_SOFT_VALUE_MAP:
            vals['lens_hard_soft'] = HARD_SOFT_VALUE_MAP[v]
        v = cell('label_print_type').lower()
        if v in LABEL_PRINT_VALUE_MAP:
            vals['label_print_type'] = LABEL_PRINT_VALUE_MAP[v]

        # SPH/CYL: chuẩn hóa định dạng "-X.YY"
        for src, field_name in [('x_sph', 'x_sph'), ('x_cyl', 'x_cyl')]:
            v = cell(src)
            if v:
                try:
                    fv = float(str(v).replace(',', '.'))
                    vals[field_name] = '%.2f' % fv if fv < 0 else '+%.2f' % fv
                except (TypeError, ValueError):
                    vals[field_name] = v

        # Boolean: Mạ polarized
        v = cell('opt_polarized').lower()
        if v in ('1', 'true', 'có', 'co', 'x', 'yes', 'y'):
            vals['opt_polarized'] = True
        elif v in ('0', 'false', 'không', 'khong', 'no', 'n'):
            vals['opt_polarized'] = False

        # Đá quý: "X kèm Y carat" hoặc "X (Y)"
        gem_raw = cell('opt_gem_combined')
        if gem_raw:
            count, carat = self._parse_gem_combined(gem_raw)
            if count is not None:
                vals['opt_gem_count'] = count
            if carat is not None:
                vals['opt_gem_carat'] = carat

        # Classification (Mã nhóm hàng) → classification_id
        code = cell('classification_code')
        if code:
            classif_id = (caches['classification_code'].get(code)
                          or caches['classification_code'].get(code.lstrip('0')))
            if classif_id:
                vals['classification_id'] = classif_id
            # categ_id: fallback lookup theo code (nếu product.category cũng dùng cùng code)
            categ_id = (caches['category_code'].get(code)
                        or caches['category_code'].get(code.lstrip('0')))
            if categ_id:
                vals['categ_id'] = categ_id

        # Country / Brand / UoM
        v = cell('country_code')
        if v and caches['country_code'].get(v.upper()):
            vals['country_id'] = caches['country_code'][v.upper()]
        v = cell('brand_code')
        if v and caches['brand_code'].get(v):
            vals['brand_id'] = caches['brand_code'][v]
        v = cell('uom_name')
        if v and caches['uom_name'].get(v):
            vals['uom_id'] = caches['uom_name'][v]
            vals['uom_po_id'] = caches['uom_name'][v]

        # Supplier
        v = cell('supplier_ref')
        if v and caches['supplier_ref'].get(v):
            vals['seller_ids'] = [(0, 0, {'partner_id': caches['supplier_ref'][v]})]

        # Taxes
        for tok, field_name, cache_key in [
            ('taxes_code', 'taxes_id', 'taxes_name'),
            ('supplier_taxes_code', 'supplier_taxes_id', 'supplier_taxes_name'),
        ]:
            ids = []
            for code in cell_csv(tok):
                tid = caches[cache_key].get(code)
                if tid:
                    ids.append(tid)
            if ids:
                vals[field_name] = [(6, 0, ids)]

        # Many2one (cid/code lookup)
        m2o_pairs = [
            ('lens_index_name', 'lens_index_id', 'lens_index_name'),
            ('lens_uv_cid', 'lens_uv_id', 'lens_uv_cid'),
            ('lens_cl_pho_cid', 'lens_cl_pho_id', 'lens_cl_pho_cid'),
            ('opt_frame_type_cid', 'opt_frame_type_id', 'opt_frame_type_cid'),
            ('opt_frame_structure_cid', 'opt_frame_structure_id', 'opt_frame_structure_cid'),
            ('opt_shape_cid', 'opt_shape_id', 'opt_shape_cid'),
            ('opt_ve_cid', 'opt_ve_id', 'opt_ve_cid'),
            ('opt_temple_tip_cid', 'opt_temple_tip_id', 'opt_temple_tip_cid'),
            ('opt_material_ve_cid', 'opt_material_ve_id', 'opt_material_cid'),
            ('opt_material_lens_cid', 'opt_material_lens_id', 'opt_material_cid'),
        ]
        for tok, field_name, cache_key in m2o_pairs:
            v = cell(tok)
            if v and caches[cache_key].get(v):
                vals[field_name] = caches[cache_key][v]

        # Many2many (csv cid/code lookup)
        m2m_pairs = [
            ('lens_design_cid', 'lens_design_ids', 'lens_design_cid'),
            ('lens_material_code', 'lens_material_ids', 'lens_material_code'),
            ('lens_film_cid', 'lens_film_ids', 'lens_film_cid'),
            ('lens_coating_cid', 'lens_coating_ids', 'lens_coating_cid'),
            ('opt_materials_front_cid', 'opt_materials_front_ids', 'opt_material_cid'),
            ('opt_materials_temple_cid', 'opt_materials_temple_ids', 'opt_material_cid'),
            ('opt_material_temple_tip_cid', 'opt_material_temple_tip_ids', 'opt_material_cid'),
            ('opt_coating_cid', 'opt_coating_ids', 'lens_coating_cid'),
        ]
        for tok, field_name, cache_key in m2m_pairs:
            ids = []
            for v in cell_csv(tok):
                rid = caches[cache_key].get(v)
                if rid:
                    ids.append(rid)
            if ids:
                vals[field_name] = [(6, 0, ids)]

        return vals

    @staticmethod
    def _parse_gem_combined(raw):
        """Parse "X kèm Y carat" / "X (Y carat)" / "X / Y" → (count, carat)."""
        if not raw:
            return None, None
        text = str(raw).lower()
        nums = re.findall(r'\d+(?:[.,]\d+)?', text)
        if not nums:
            return None, None
        count = None
        carat = None
        try:
            count = int(float(nums[0].replace(',', '.')))
        except (TypeError, ValueError):
            count = None
        if len(nums) > 1:
            try:
                carat = float(nums[1].replace(',', '.'))
            except (TypeError, ValueError):
                carat = None
        return count, carat

    # ────────────────────────────────────────────────────────────
    #   VN-LABEL: PREVIEW HTML (3 nhóm)
    # ────────────────────────────────────────────────────────────

    _LEVEL_STYLES = {
        'error': ('danger', '#dc3545'),
        'warn': ('warning', '#856404'),
        'ok': ('success', '#155724'),
    }

    _CLASSIF_META = {
        'lens': ('🔬 Tròng kính', 'primary', '#0d6efd'),
        'frame': ('👓 Gọng kính', 'success', '#198754'),
        'accessory': ('🧰 Phụ kiện', 'info', '#0dcaf0'),
        'other': ('🗂️ Khác / chưa phân loại', 'secondary', '#6c757d'),
    }

    def _build_vn_preview_html(self, header, rows, issues, buckets):
        parts = []

        total = sum(len(v) for v in buckets.values())
        parts.append(
            '<div class="card mb-3"><div class="card-header fw-bold">📊 Tổng quan</div>'
            '<div class="card-body">'
            '<div>Tổng số dòng dữ liệu: <strong>%s</strong></div>'
            '</div></div>' % total
        )

        # Issues
        if issues:
            rows_html = []
            for issue in issues:
                badge, color = self._LEVEL_STYLES.get(
                    issue['level'], ('secondary', '#6c757d'))
                title = '<strong style="color:%s;">%s</strong>' % (color, issue['title'])
                detail_items = ''.join(
                    '<li>%s</li>' % d for d in issue.get('details', []))
                detail_html = '<ul class="mb-0 ps-3">%s</ul>' % detail_items if detail_items else ''
                rows_html.append(
                    '<div class="border-start border-3 ps-3 mb-2" style="border-color:%s !important;">'
                    '%s%s</div>' % (color, title, detail_html)
                )
            parts.append(
                '<div class="card mb-3">'
                '<div class="card-header fw-bold">🔍 Chi tiết kiểm tra</div>'
                '<div class="card-body">%s</div></div>' % ''.join(rows_html)
            )

        # Verdict
        has_errors = any(i['level'] == 'error' for i in issues)
        has_warns = any(i['level'] == 'warn' for i in issues)
        if has_errors:
            badge = 'danger'
            err_n = sum(1 for i in issues if i['level'] == 'error')
            msg = _('Phát hiện %s lỗi nghiêm trọng. Sửa file rồi thử lại.') % err_n
        elif has_warns:
            badge = 'warning'
            msg = _('Có %s cảnh báo. Vẫn có thể import; các giá trị thiếu sẽ bỏ qua.') % len(issues)
        else:
            badge = 'success'
            msg = _('File hợp lệ, sẵn sàng import.')
        parts.append(
            '<div class="alert alert-%s mb-0"><strong>%s</strong> %s</div>'
            % (badge, _('Kết quả:'), msg)
        )

        return '<div>%s</div>' % ''.join(parts)

    # ────────────────────────────────────────────────────────────
    #   TECHNICAL FORMAT (hành vi cũ, giữ nguyên)
    # ────────────────────────────────────────────────────────────

    def _action_test_technical(self, raw):
        header, rows, errors = self._parse_excel(raw)
        if errors:
            self.write({
                'state': 'preview',
                'preview_text': self._html_error_panel(
                    _('Lỗi khi đọc file'), '<br/>'.join(errors)),
            })
            return self._reopen()
        if not rows:
            self.write({
                'state': 'preview',
                'preview_text': self._html_error_panel(
                    _('Lỗi'), _('File không có dữ liệu.')),
            })
            return self._reopen()

        product_fields = self.env['product.template']._fields
        virtual_columns = set(VNOP_TEMPLATE_VIRTUAL_IMPORT_COLUMNS)
        col_map = {h: i for i, h in enumerate(header) if h}
        issues = []

        unknown = [
            h for h in header
            if h and h not in product_fields and h not in virtual_columns and '/' not in h
        ]
        if unknown:
            issues.append({
                'level': 'warn',
                'title': _('Cột không nhận dạng được'),
                'details': [', '.join(unknown)],
            })

        self._test_required_fields(issues, rows, col_map)
        self._test_duplicate_names(issues, rows, col_map)
        self._test_existing_names(issues, rows, col_map)
        self._test_duplicate_barcodes(issues, rows, col_map)
        self._test_existing_barcodes(issues, rows, col_map)
        self._test_relational_fields(issues, rows, col_map, product_fields)

        html = self._build_preview_html(header, rows, col_map, issues)
        self.write({'state': 'preview', 'preview_text': html})
        return self._reopen()

    def _action_import_technical(self, raw):
        header, rows, errors = self._parse_excel(raw)
        if errors:
            self.write({
                'state': 'done',
                'imported_count': 0,
                'error_text': '\n'.join(errors),
            })
            return self._reopen()
        if not rows:
            self.write({
                'state': 'done',
                'imported_count': 0,
                'error_text': _('File không có dữ liệu.'),
            })
            return self._reopen()

        # Chặn import nếu trùng tên/barcode (trong file hoặc DB).
        col_map = {h: i for i, h in enumerate(header) if h}
        dup_errors = self._assert_no_duplicates(rows, col_map)
        if dup_errors:
            self.write({
                'state': 'done', 'imported_count': 0,
                'error_text': _('Không thể import do trùng tên/mã vạch:') + '\n\n' + '\n\n'.join(dup_errors),
            })
            return self._reopen()

        ProductTemplate = self.env['product.template'].with_context(
            import_file=True,
            vnop_import_product_type=self.product_type if self.product_type != 'auto' else 'lens',
        )
        try:
            result = ProductTemplate.load(header, rows)
        except Exception as e:
            self.write({
                'state': 'done',
                'imported_count': 0,
                'error_text': str(e),
            })
            return self._reopen()

        messages = result.get('messages', [])
        error_messages = [m.get('message', '') for m in messages if m.get('type') == 'error']
        if error_messages:
            self.write({
                'state': 'done',
                'imported_count': 0,
                'error_text': '\n'.join(error_messages),
            })
            return self._reopen()

        ids = result.get('ids') or []
        created = self.env['product.template'].browse(ids) if ids else self.env['product.template']
        lens_n = sum(1 for p in created if p.classification_type == 'lens')
        frame_n = sum(1 for p in created if p.classification_type == 'frame')
        acc_n = sum(1 for p in created if p.classification_type == 'accessory')
        other_n = len(created) - lens_n - frame_n - acc_n

        self.write({
            'state': 'done',
            'imported_count': len(ids),
            'imported_lens_count': lens_n,
            'imported_frame_count': frame_n,
            'imported_accessory_count': acc_n,
            'imported_other_count': other_n,
            'imported_product_ids': [(6, 0, ids)],
            'error_text': False,
        })
        return self._reopen()

    _DUP_SAMPLE_LIMIT = 10

    def _assert_no_duplicates(self, rows, col_map):
        """Phát hiện trùng tên/barcode trong file và đã tồn tại trong DB.
        Trả về list message lỗi (rỗng = OK). Caller phải dừng import nếu non-empty.
        """
        name_idx = col_map.get('name')
        bc_idx = col_map.get('barcode')
        if name_idx is None and bc_idx is None:
            return []

        def _cell(row, idx):
            return (row[idx] or '').strip() if idx is not None and idx < len(row) else ''

        # Trùng trong file
        seen_names, dup_names_in_file = {}, {}
        seen_bcs, dup_bcs_in_file = {}, {}
        all_names, all_bcs = [], []
        for i, row in enumerate(rows):
            nm = _cell(row, name_idx)
            bc = _cell(row, bc_idx)
            if nm:
                all_names.append(nm)
                if nm in seen_names:
                    dup_names_in_file.setdefault(nm, [seen_names[nm]]).append(i + 1)
                else:
                    seen_names[nm] = i + 1
            if bc:
                all_bcs.append(bc)
                if bc in seen_bcs:
                    dup_bcs_in_file.setdefault(bc, [seen_bcs[bc]]).append(i + 1)
                else:
                    seen_bcs[bc] = i + 1

        # Trùng trong DB
        Tmpl = self.env['product.template'].with_context(active_test=False)
        existing_names = set()
        if all_names:
            existing_names = set(Tmpl.search(
                [('name', 'in', list(set(all_names)))]
            ).mapped('name'))
        existing_bcs = set()
        if all_bcs:
            existing_bcs = set(Tmpl.search(
                [('barcode', 'in', list(set(all_bcs)))]
            ).mapped('barcode'))

        errors = []
        if dup_names_in_file:
            sample = list(dup_names_in_file.items())[:self._DUP_SAMPLE_LIMIT]
            details = ['  - "%s" (dòng: %s)' % (nm, ', '.join(str(r) for r in rs))
                       for nm, rs in sample]
            extra = len(dup_names_in_file) - len(sample)
            if extra > 0:
                details.append(_('  ... và %s tên khác') % extra)
            errors.append(_('Tên trùng trong file (%s tên):\n%s') % (
                len(dup_names_in_file), '\n'.join(details),
            ))
        if dup_bcs_in_file:
            sample = list(dup_bcs_in_file.items())[:self._DUP_SAMPLE_LIMIT]
            details = ['  - "%s" (dòng: %s)' % (bc, ', '.join(str(r) for r in rs))
                       for bc, rs in sample]
            extra = len(dup_bcs_in_file) - len(sample)
            if extra > 0:
                details.append(_('  ... và %s mã khác') % extra)
            errors.append(_('Mã vạch trùng trong file (%s mã):\n%s') % (
                len(dup_bcs_in_file), '\n'.join(details),
            ))
        if existing_names:
            sample = list(existing_names)[:self._DUP_SAMPLE_LIMIT]
            extra = len(existing_names) - len(sample)
            tail = (_('\n  ... và %s tên khác') % extra) if extra > 0 else ''
            errors.append(_('Tên đã tồn tại trong hệ thống (%s tên):\n  - %s%s') % (
                len(existing_names), '\n  - '.join(sample), tail,
            ))
        if existing_bcs:
            sample = list(existing_bcs)[:self._DUP_SAMPLE_LIMIT]
            extra = len(existing_bcs) - len(sample)
            tail = (_('\n  ... và %s mã khác') % extra) if extra > 0 else ''
            errors.append(_('Mã vạch đã tồn tại trong hệ thống (%s mã):\n  - %s%s') % (
                len(existing_bcs), '\n  - '.join(sample), tail,
            ))
        return errors

    # ────────────────────────────────────────────────────────────
    #   BUILDERS / VALIDATION (existing technical-format helpers)
    # ────────────────────────────────────────────────────────────

    def _build_preview_html(self, header, rows, col_map, issues):
        parts = []
        col_tags = ''.join(
            '<span class="badge text-bg-secondary me-1 mb-1">%s</span>' % h
            for h in header if h
        )
        lbl_overview = _('Tổng quan')
        lbl_cols = _('Số cột')
        lbl_rows = _('Số dòng dữ liệu')
        parts.append(
            '<div class="card mb-3">'
            '<div class="card-header fw-bold">%s</div>'
            '<div class="card-body">'
            '<table class="table table-sm mb-2" style="width:auto;">'
            '<tr><td class="text-muted pe-3">%s</td><td class="fw-bold">%s</td></tr>'
            '<tr><td class="text-muted pe-3">%s</td><td class="fw-bold">%s</td></tr>'
            '</table>'
            '<div>%s</div>'
            '</div></div>' % (lbl_overview, lbl_cols, len(header),
                              lbl_rows, len(rows), col_tags)
        )

        if issues:
            rows_html = []
            for issue in issues:
                level_style = self._LEVEL_STYLES.get(
                    issue['level'], ('secondary', '#6c757d'))
                color = level_style[1]
                title = '<strong style="color:%s;">%s</strong>' % (color, issue['title'])
                detail_items = ''.join(
                    '<li>%s</li>' % d for d in issue.get('details', []))
                detail_html = '<ul class="mb-0 ps-3">%s</ul>' % detail_items if detail_items else ''
                rows_html.append(
                    '<div class="border-start border-3 ps-3 mb-2" style="border-color:%s !important;">'
                    '%s%s</div>' % (color, title, detail_html)
                )
            lbl_detail = _('Chi tiết kiểm tra')
            parts.append(
                '<div class="card mb-3">'
                '<div class="card-header fw-bold">%s</div>'
                '<div class="card-body">%s</div></div>' % (lbl_detail, ''.join(rows_html))
            )

        has_errors = any(i['level'] == 'error' for i in issues)
        has_warns = any(i['level'] == 'warn' for i in issues)
        if has_errors:
            badge = 'danger'
            error_count = sum(1 for i in issues if i['level'] == 'error')
            msg = _('Phát hiện %s lỗi. Không thể import.') % error_count
        elif has_warns:
            badge = 'warning'
            msg = _('Phát hiện %s cảnh báo. Vui lòng kiểm tra trước khi import.') % len(issues)
        else:
            badge = 'success'
            msg = _('Không phát hiện lỗi. Sẵn sàng import.')

        lbl_result = _('Kết quả:')
        parts.append(
            '<div class="alert alert-%s mb-0" role="alert">'
            '<strong>%s</strong> %s</div>' % (badge, lbl_result, msg)
        )

        return '<div>%s</div>' % ''.join(parts)

    @staticmethod
    def _html_error_panel(title, body):
        return (
            '<div class="alert alert-danger" role="alert">'
            '<strong>%s</strong><br/>%s</div>' % (title, body)
        )

    def _test_required_fields(self, issues, rows, col_map):
        required = ['name']
        for field_name in required:
            if field_name not in col_map:
                issues.append({
                    'level': 'error',
                    'title': _('Thiếu cột bắt buộc: %s') % field_name,
                })
                continue
            idx = col_map[field_name]
            empty_rows = [i + 1 for i, row in enumerate(rows) if not row[idx].strip()]
            if empty_rows:
                sample = empty_rows[:self._MAX_SAMPLE]
                extra = len(empty_rows) - self._MAX_SAMPLE
                detail = _('Dòng: %s') % ', '.join(str(r) for r in sample)
                if extra > 0:
                    detail += _(' ... và %s dòng khác') % extra
                issues.append({
                    'level': 'error',
                    'title': _('%s trống ở %s dòng') % (field_name, len(empty_rows)),
                    'details': [detail],
                })

    def _test_duplicate_names(self, issues, rows, col_map):
        if 'name' not in col_map:
            return
        idx = col_map['name']
        seen = {}
        duplicates = {}
        for i, row in enumerate(rows):
            name = row[idx].strip()
            if not name:
                continue
            if name in seen:
                duplicates.setdefault(name, [seen[name]]).append(i + 1)
            else:
                seen[name] = i + 1

        if not duplicates:
            return

        details = []
        sample = list(duplicates.keys())[:self._MAX_SAMPLE]
        for name in sample:
            row_nums = ', '.join(str(r) for r in duplicates[name][:self._MAX_SAMPLE])
            details.append('"%s" (dòng: %s)' % (name, row_nums))
        if len(duplicates) > self._MAX_SAMPLE:
            details.append(_('... và %s tên khác') % (len(duplicates) - self._MAX_SAMPLE))
        issues.append({
            'level': 'error',
            'title': _('Tên trùng trong file: %s tên') % len(duplicates),
            'details': details,
        })

    def _test_duplicate_barcodes(self, issues, rows, col_map):
        if 'barcode' not in col_map:
            return
        idx = col_map['barcode']
        seen = {}
        duplicates = {}
        for i, row in enumerate(rows):
            bc = (row[idx] or '').strip() if idx < len(row) else ''
            if not bc:
                continue
            if bc in seen:
                duplicates.setdefault(bc, [seen[bc]]).append(i + 1)
            else:
                seen[bc] = i + 1
        if not duplicates:
            return
        details = []
        sample = list(duplicates.keys())[:self._MAX_SAMPLE]
        for bc in sample:
            row_nums = ', '.join(str(r) for r in duplicates[bc][:self._MAX_SAMPLE])
            details.append('"%s" (dòng: %s)' % (bc, row_nums))
        if len(duplicates) > self._MAX_SAMPLE:
            details.append(_('... và %s mã khác') % (len(duplicates) - self._MAX_SAMPLE))
        issues.append({
            'level': 'error',
            'title': _('Mã vạch trùng trong file: %s mã') % len(duplicates),
            'details': details,
        })

    def _test_existing_barcodes(self, issues, rows, col_map):
        if 'barcode' not in col_map:
            return
        idx = col_map['barcode']
        unique_bcs = list({(row[idx] or '').strip() for row in rows
                           if idx < len(row) and (row[idx] or '').strip()})
        if not unique_bcs:
            return
        existing = self.env['product.template'].with_context(active_test=False).search_read(
            [('barcode', 'in', unique_bcs)], ['barcode'],
        )
        existing_bcs = [r['barcode'] for r in existing if r['barcode']]
        if not existing_bcs:
            return
        details = [', '.join(existing_bcs[:10])]
        if len(existing_bcs) > 10:
            details.append(_('... và %s mã khác') % (len(existing_bcs) - 10))
        issues.append({
            'level': 'error',
            'title': _('Mã vạch đã tồn tại trong hệ thống: %s mã') % len(existing_bcs),
            'details': details,
        })

    def _test_existing_names(self, issues, rows, col_map):
        if 'name' not in col_map:
            return
        idx = col_map['name']
        unique_names = list({row[idx].strip() for row in rows if row[idx].strip()})
        if not unique_names:
            return

        existing = self.env['product.template'].with_context(active_test=False).search_read(
            [('name', 'in', unique_names)],
            ['name'],
        )
        existing_names = [r['name'] for r in existing if r['name']]
        if not existing_names:
            return

        details = []
        sample = existing_names[:10]
        details.append(', '.join(sample))
        if len(existing_names) > 10:
            details.append(_('... và %s tên khác') % (len(existing_names) - 10))
        issues.append({
            'level': 'error',
            'title': _('Tên đã tồn tại trong hệ thống: %s tên') % len(existing_names),
            'details': details,
        })

    def _test_relational_fields(self, issues, rows, col_map, product_fields):
        for field_name, idx in col_map.items():
            if '/' in field_name or field_name not in product_fields:
                continue
            field = product_fields[field_name]
            if field.type != 'many2one':
                continue

            raw_values = {row[idx].strip() for row in rows if row[idx].strip()}
            if not raw_values:
                continue

            values = raw_values

            lookup = self._FIELD_LOOKUP_CONFIG.get(field_name)
            if lookup:
                search_field, domain = lookup
            else:
                CoModel = self.env[field.comodel_name]
                search_field = CoModel._rec_name or 'name'
                domain = []

            CoModel = self.env[field.comodel_name]
            if search_field not in CoModel._fields:
                continue

            found = CoModel.search_read(
                domain + [(search_field, 'in', list(values))],
                [search_field],
            )
            found_values = {r[search_field] for r in found if r.get(search_field)}
            not_found = values - found_values
            if not not_found:
                continue

            display_code = VNOP_TEMPLATE_FIELD_CODE_OVERRIDES.get(
                field_name, field_name)
            details = []
            sample = list(not_found)[:self._MAX_SAMPLE]
            details.append(', '.join('"%s"' % v for v in sample))
            if len(not_found) > self._MAX_SAMPLE:
                details.append(_('... và %s giá trị khác') % (len(not_found) - self._MAX_SAMPLE))
            issues.append({
                'level': 'warn',
                'title': _('%s: %s giá trị không tìm thấy trong %s') % (
                    display_code, len(not_found), field.comodel_name),
                'details': details,
            })

    def _parse_excel(self, file_bytes):
        errors = []
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            return [], [], [_('Không đọc được file Excel: %s') % str(e)]

        ws = wb.active
        if ws is None:
            wb.close()
            return [], [], [_('File Excel không có sheet nào.')]

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([self._normalize_cell(c) for c in row])
        wb.close()

        if not all_rows:
            return [], [], [_('File Excel trống.')]

        product_fields = set(self.env['product.template']._fields)
        header_index = self._find_header_index(all_rows, product_fields)

        if header_index is None:
            return [], [], [_('Không tìm thấy dòng header chứa tên trường kỹ thuật (vd: name, categ_code, brand_code...).')]

        header = [self._translate_token(self._normalize_cell(c))
                  for c in all_rows[header_index]]

        data_start = self._find_data_start(all_rows, header_index, header, product_fields)

        header_len = len(header)
        data_rows = []
        for row in all_rows[data_start:]:
            if all(not self._normalize_cell(c) for c in row):
                continue
            padded = list(row) + [''] * max(0, header_len - len(row))
            data_rows.append(padded[:header_len])

        if not data_rows:
            errors.append(_('File không có dữ liệu (chỉ có header).'))

        return header, data_rows, errors

    def _translate_token(self, token):
        return self._CODE_TO_FIELD.get(token, token)

    def _find_header_index(self, rows, product_fields):
        marker_fields = {
            'categ_id', 'uom_id', 'brand_id',
            'len_type', 'opt_sku', 'opt_model',
            'design_id', 'shape_id', 'material_id', 'color_id',
        }
        code_overrides = VNOP_TEMPLATE_FIELD_CODE_OVERRIDES
        marker_display = {code_overrides.get(f, f) for f in marker_fields
                          if f in code_overrides}
        all_markers = marker_fields | marker_display
        all_known = product_fields | set(self._CODE_TO_FIELD.keys())

        best_index = None
        best_score = -1

        for index, row in enumerate(rows):
            tokens = {self._normalize_cell(c) for c in row if self._normalize_cell(c)}
            matched = tokens & all_known
            if not matched:
                continue
            score = len(matched)
            has_name = 'name' in matched
            has_marker = bool(matched & all_markers)
            if has_name and has_marker and score > best_score:
                best_index = index
                best_score = score

        if best_index is not None:
            return best_index

        for index, row in enumerate(rows):
            tokens = {self._normalize_cell(c) for c in row if self._normalize_cell(c)}
            score = len(tokens & all_known)
            if score >= 3 and score > best_score:
                best_index = index
                best_score = score

        return best_index

    def _find_data_start(self, rows, header_index, header_tokens, product_fields):
        key_fields = {
            'name', 'default_code', 'standard_price', 'list_price',
            'len_type', 'opt_model', 'opt_sku',
        }
        key_display = {
            'categ_code', 'uom_name', 'brand_code',
            'categ_id', 'uom_id', 'brand_id',
            'design_cid', 'shape_cid', 'design_id', 'shape_id',
        }
        key_fields = key_fields | key_display
        key_indices = [i for i, t in enumerate(header_tokens) if t in key_fields]
        header_len = len(header_tokens)

        for row_index in range(header_index + 1, len(rows)):
            row = rows[row_index]
            non_empty = [
                i for i in range(min(len(row), header_len))
                if self._normalize_cell(row[i])
            ]
            if not non_empty:
                continue
            if key_indices and any(i in non_empty for i in key_indices):
                return row_index
            if len(non_empty) >= 2:
                return row_index

        return len(rows)

    @staticmethod
    def _normalize_cell(value):
        if value in (None, False):
            return ''
        text = str(value).strip()
        if not text or text.lower() in ('none', 'null', 'nan'):
            return ''
        return text

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
