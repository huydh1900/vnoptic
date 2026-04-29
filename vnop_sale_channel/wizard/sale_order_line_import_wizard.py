# -*- coding: utf-8 -*-
import base64
import io
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from odoo import fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Template columns: (technical_key, label, hint)
TEMPLATE_COLUMNS = [
    ("default_code", "Sản phẩm (Mã SP)", "Điền mã nội bộ (default_code), required"),
    ("product_uom_qty", "Số lượng", "Số, required"),
    ("price_unit", "Đơn giá", "Số, required"),
    ("taxes", "Thuế", "VD: 8% hoặc 10%"),
]

BASE_FONT_NAME = "Times New Roman"
BASE_FONT_SIZE = 14
HEADER_FONT = Font(name=BASE_FONT_NAME, bold=True, color="FFFFFF", size=BASE_FONT_SIZE)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HINT_FONT = Font(name=BASE_FONT_NAME, italic=True, color="808080", size=BASE_FONT_SIZE)
HINT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LABEL_FONT = Font(name=BASE_FONT_NAME, bold=True, size=BASE_FONT_SIZE)
DATA_ALIGNMENT = Alignment(horizontal="left")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class SaleOrderLineImportWizard(models.TransientModel):
    _name = "sale.order.line.import.wizard"
    _description = "Import dòng đơn bán từ Excel"

    order_id = fields.Many2one("sale.order", required=True, readonly=True)
    file_data = fields.Binary(string="File Excel")
    file_name = fields.Char()

    state = fields.Selection([
        ("upload", "Upload"),
        ("validated", "Kiểm thử"),
        ("done", "Hoàn tất"),
    ], default="upload")
    validated_count = fields.Integer(string="Dòng hợp lệ", readonly=True)
    imported_count = fields.Integer(readonly=True)
    error_text = fields.Text(readonly=True)

    def action_validate(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Vui lòng upload file Excel."))

        raw = base64.b64decode(self.file_data)
        rows, errors = self._parse_excel(raw)

        if errors:
            self.write({
                "state": "validated",
                "validated_count": 0,
                "error_text": "\n".join(errors),
            })
            return self._reopen()

        line_vals_list, import_errors = self._prepare_line_vals(rows)

        self.write({
            "state": "validated",
            "validated_count": len(line_vals_list),
            "error_text": "\n".join(import_errors) if import_errors else False,
        })
        return self._reopen()

    def action_import(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Vui lòng upload file Excel."))

        raw = base64.b64decode(self.file_data)
        rows, errors = self._parse_excel(raw)

        if errors:
            self.write({
                "state": "done",
                "imported_count": 0,
                "error_text": "\n".join(errors),
            })
            return self._reopen()

        line_vals_list, import_errors = self._prepare_line_vals(rows)

        if import_errors:
            self.write({
                "state": "done",
                "imported_count": 0,
                "error_text": "\n".join(import_errors),
            })
            return self._reopen()

        if not line_vals_list:
            raise UserError(_("File không có dữ liệu hợp lệ."))

        self.env["sale.order.line"].create(line_vals_list)
        self.write({
            "state": "done",
            "imported_count": len(line_vals_list),
        })
        # Notification cảnh báo tồn kho không return từ đây nữa: form sale.order
        # phía sau wizard refresh order_line → useEffect bên JS sẽ bắt thay đổi
        # và tự gọi check_stock_warning. JS có module-level singleton đảm bảo
        # chỉ 1 notification hiển thị cùng lúc.
        return self._reopen()

    def action_back_upload(self):
        self.ensure_one()
        self.write({
            "state": "upload",
            "validated_count": 0,
            "imported_count": 0,
            "error_text": False,
        })
        return self._reopen()

    def _parse_excel(self, file_bytes):
        errors = []
        rows = []

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            return [], [_("Không đọc được file Excel: %s") % str(e)]

        ws = wb.active
        if ws is None:
            wb.close()
            return [], [_("File Excel không có sheet nào.")]

        header_row = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
            header_row.append(str(cell.value).strip() if cell.value else "")

        col_keys = [c[0] for c in TEMPLATE_COLUMNS]
        col_mapping = {}
        for col_idx, val in enumerate(header_row):
            if val in col_keys:
                col_mapping[val] = col_idx

        missing = [k for k in ("default_code", "product_uom_qty", "price_unit") if k not in col_mapping]
        if missing:
            wb.close()
            return [], [_("Thiếu cột bắt buộc: %s") % ", ".join(missing)]

        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if row_cells is None:
                continue
            if all(c is None or str(c).strip() == "" for c in row_cells):
                continue

            row_data = {"_row": row_idx}
            for key, ci in col_mapping.items():
                row_data[key] = row_cells[ci] if ci < len(row_cells) else None
            rows.append(row_data)

        wb.close()

        if not rows:
            errors.append(_("File không có dữ liệu (data bắt đầu từ row 3)."))

        return rows, errors

    def _prepare_line_vals(self, rows):
        errors = []
        vals_list = []

        codes = [self._normalize_code(r.get("default_code")) for r in rows if r.get("default_code")]
        products = self.env["product.product"].search([("default_code", "in", codes)])
        product_by_code = {p.default_code: p for p in products if p.default_code}

        sale_taxes = self.env["account.tax"].search([("type_tax_use", "=", "sale")])
        tax_by_amount = {tax.amount: tax for tax in sale_taxes}

        for row in rows:
            row_num = row["_row"]

            code = self._normalize_code(row.get("default_code"))
            if not code:
                errors.append(_("Dòng %d: Thiếu mã sản phẩm.") % row_num)
                continue
            product = product_by_code.get(code)
            if not product:
                errors.append(_("Dòng %d: Không tìm thấy sản phẩm với mã '%s'.") % (row_num, code))
                continue

            try:
                qty = float(row.get("product_uom_qty") or 0)
                if qty <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(_("Dòng %d: Số lượng không hợp lệ.") % row_num)
                continue

            try:
                price = float(row.get("price_unit") or 0)
                if price < 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(_("Dòng %d: Đơn giá không hợp lệ.") % row_num)
                continue

            tax_ids = []
            raw_tax = row.get("taxes")
            tax_raw = str(raw_tax or "").strip()
            if tax_raw:
                for part in tax_raw.split(","):
                    part = part.strip().replace("%", "")
                    if not part:
                        continue
                    try:
                        amount = float(part)
                    except ValueError:
                        errors.append(_("Dòng %d: Thuế '%s' không hợp lệ.") % (row_num, part))
                        continue
                    if amount < 1 and isinstance(raw_tax, float):
                        amount = round(amount * 100, 2)
                    tax = tax_by_amount.get(amount)
                    if not tax:
                        errors.append(_("Dòng %d: Không tìm thấy thuế bán hàng %s%%.") % (row_num, part))
                        continue
                    tax_ids.append(tax.id)

            vals_list.append({
                "order_id": self.order_id.id,
                "product_id": product.id,
                "product_uom_qty": qty,
                "price_unit": price,
                "tax_id": [(6, 0, tax_ids)],
                "name": product.display_name,
            })

        return vals_list, errors

    @staticmethod
    def _normalize_code(raw):
        if raw is None:
            return ""
        if isinstance(raw, float):
            if raw == int(raw):
                return str(int(raw)).strip()
            return str(raw).strip()
        return str(raw).strip()

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    @staticmethod
    def generate_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        text_columns = {"default_code", "taxes"}

        for col_idx, (key, label, hint) in enumerate(TEMPLATE_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)

            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

            cell_label = ws.cell(row=2, column=col_idx, value=label)
            cell_label.font = LABEL_FONT
            cell_label.alignment = DATA_ALIGNMENT
            cell_label.border = THIN_BORDER
            if "required" in hint.lower():
                cell_label.fill = REQUIRED_FILL

            # Excel column width ~ ký tự của font mặc định 11pt; font 14 rộng hơn ~1.3x
            # nên scale lên để label (vd "Sản phẩm (Mã SP)") không bị cắt.
            max_len = max(len(key), len(label))
            ws.column_dimensions[col_letter].width = max(max_len * 1.4 + 4, 18)

        text_col_indices = [
            i for i, (k, _l, _h) in enumerate(TEMPLATE_COLUMNS, 1) if k in text_columns
        ]
        for row in range(3, 1003):
            for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
                c = ws.cell(row=row, column=col_idx)
                c.font = Font(name=BASE_FONT_NAME, size=BASE_FONT_SIZE)
                c.alignment = DATA_ALIGNMENT
                if col_idx in text_col_indices:
                    c.number_format = numbers.FORMAT_TEXT

        ws.freeze_panes = "A3"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
