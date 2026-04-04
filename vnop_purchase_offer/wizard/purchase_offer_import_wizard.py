# -*- coding: utf-8 -*-
import base64
import io
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from odoo import fields, models, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Template columns: (technical_key, label, hint)
TEMPLATE_COLUMNS = [
    ("default_code", "Sản phẩm (Mã SP)", "Điền mã nội bộ (default_code), required"),
    ("quantity", "SL dự kiến", "Số, required"),
    ("expected_price", "Giá dự kiến", "Số, required"),
    ("taxes", "Thuế", "VD: 8% hoặc 10%"),
    ("description", "Mô tả", "Text"),
]

BASE_FONT_NAME = "Times New Roman"
HEADER_FONT = Font(name=BASE_FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HINT_FONT = Font(name=BASE_FONT_NAME, italic=True, color="808080", size=9)
HINT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LABEL_FONT = Font(name=BASE_FONT_NAME, bold=True, size=10)
DATA_ALIGNMENT = Alignment(horizontal="left")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class PurchaseOfferImportWizard(models.TransientModel):
    _name = "purchase.offer.import.wizard"
    _description = "Import dòng đề nghị mua hàng từ Excel"

    offer_id = fields.Many2one("purchase.offer", required=True, readonly=True)
    file_data = fields.Binary(string="File Excel")
    file_name = fields.Char()

    # Results
    state = fields.Selection([
        ("upload", "Upload"),
        ("validated", "Kiểm thử"),
        ("done", "Hoàn tất"),
    ], default="upload")
    validated_count = fields.Integer(string="Dòng hợp lệ", readonly=True)
    imported_count = fields.Integer(readonly=True)
    error_text = fields.Text(readonly=True)

    def action_validate(self):
        """Kiểm thử file Excel — chỉ validate, không ghi DB."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Vui lòng upload file Excel."))

        raw = base64.b64decode(self.file_data)
        rows, errors = self._parse_excel(raw)

        if errors:
            self.write({
                "state": "validated",
                "validated_count": 0,
                "error_text": "\n".join(errors),
            })
            return self._reopen()

        line_vals_list, import_errors = self._prepare_line_vals(rows)

        self.write({
            "state": "validated",
            "validated_count": len(line_vals_list),
            "error_text": "\n".join(import_errors) if import_errors else False,
        })
        return self._reopen()

    def action_import(self):
        """Import chính thức — tạo purchase.offer.line."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Vui lòng upload file Excel."))

        raw = base64.b64decode(self.file_data)
        rows, errors = self._parse_excel(raw)

        if errors:
            self.write({
                "state": "done",
                "imported_count": 0,
                "error_text": "\n".join(errors),
            })
            return self._reopen()

        line_vals_list, import_errors = self._prepare_line_vals(rows)

        if import_errors:
            self.write({
                "state": "done",
                "imported_count": 0,
                "error_text": "\n".join(import_errors),
            })
            return self._reopen()

        if not line_vals_list:
            raise UserError(_("File không có dữ liệu hợp lệ."))

        self.env["purchase.offer.line"].create(line_vals_list)
        self.write({
            "state": "done",
            "imported_count": len(line_vals_list),
        })
        return self._reopen()

    def action_back_upload(self):
        """Quay lại bước upload."""
        self.ensure_one()
        self.write({
            "state": "upload",
            "validated_count": 0,
            "imported_count": 0,
            "error_text": False,
        })
        return self._reopen()

    def _parse_excel(self, file_bytes):
        """Parse Excel → list[dict] với keys: default_code, quantity, expected_price, taxes, description."""
        errors = []
        rows = []

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            return [], [_("Không đọc được file Excel: %s") % str(e)]

        ws = wb.active
        if ws is None:
            wb.close()
            return [], [_("File Excel không có sheet nào.")]

        # Row 1 = technical key, row 2 = label, row 3 = hint, data từ row 4
        header_row = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
            header_row.append(str(cell.value).strip() if cell.value else "")

        col_keys = [c[0] for c in TEMPLATE_COLUMNS]
        col_mapping = {}
        for col_idx, val in enumerate(header_row):
            if val in col_keys:
                col_mapping[val] = col_idx

        missing = [k for k in ("default_code", "quantity", "expected_price") if k not in col_mapping]
        if missing:
            wb.close()
            return [], [_("Thiếu cột bắt buộc: %s") % ", ".join(missing)]

        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if row_cells is None:
                continue
            if all(c is None or str(c).strip() == "" for c in row_cells):
                continue

            row_data = {"_row": row_idx}
            for key, ci in col_mapping.items():
                row_data[key] = row_cells[ci] if ci < len(row_cells) else None
            rows.append(row_data)

        wb.close()

        if not rows:
            errors.append(_("File không có dữ liệu (data bắt đầu từ row 4)."))

        return rows, errors

    def _prepare_line_vals(self, rows):
        """Chuyển parsed rows → list create vals cho purchase.offer.line."""
        errors = []
        vals_list = []

        # Pre-cache products by default_code
        codes = [self._normalize_code(r.get("default_code")) for r in rows if r.get("default_code")]
        products = self.env["product.product"].search([("default_code", "in", codes)])
        product_by_code = {p.default_code: p for p in products if p.default_code}

        # Pre-cache taxes by amount (purchase taxes only)
        purchase_taxes = self.env["account.tax"].search([("type_tax_use", "=", "purchase")])
        tax_by_amount = {tax.amount: tax for tax in purchase_taxes}

        for row in rows:
            row_num = row["_row"]

            # --- Product ---
            code = self._normalize_code(row.get("default_code"))
            if not code:
                errors.append(_("Dòng %d: Thiếu mã sản phẩm.") % row_num)
                continue
            product = product_by_code.get(code)
            if not product:
                errors.append(_("Dòng %d: Không tìm thấy sản phẩm với mã '%s'.") % (row_num, code))
                continue

            # --- Quantity ---
            try:
                qty = float(row.get("quantity") or 0)
                if qty <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(_("Dòng %d: SL dự kiến không hợp lệ.") % row_num)
                continue

            # --- Expected price ---
            try:
                price = float(row.get("expected_price") or 0)
                if price < 0:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(_("Dòng %d: Giá dự kiến không hợp lệ.") % row_num)
                continue

            # --- UoM (auto từ product) ---
            uom = product.uom_po_id or product.uom_id

            # --- Taxes ---
            tax_ids = []
            raw_tax = row.get("taxes")
            tax_raw = str(raw_tax or "").strip()
            if tax_raw:
                for part in tax_raw.split(","):
                    part = part.strip().replace("%", "")
                    if not part:
                        continue
                    try:
                        amount = float(part)
                    except ValueError:
                        errors.append(_("Dòng %d: Thuế '%s' không hợp lệ.") % (row_num, part))
                        continue
                    # Excel có thể đọc "8%" thành 0.08 → nhân 100
                    if amount < 1 and isinstance(raw_tax, float):
                        amount = round(amount * 100, 2)
                    tax = tax_by_amount.get(amount)
                    if not tax:
                        errors.append(_("Dòng %d: Không tìm thấy thuế mua hàng %s%%.") % (row_num, part))
                        continue
                    tax_ids.append(tax.id)

            # --- Description ---
            description = str(row.get("description") or "").strip() or product.display_name

            vals_list.append({
                "offer_id": self.offer_id.id,
                "product_id": product.id,
                "uom_id": uom.id,
                "quantity": qty,
                "expected_price": price,
                "taxes_id": [(6, 0, tax_ids)],
                "description": description,
            })

        return vals_list, errors

    @staticmethod
    def _normalize_code(raw):
        """Chuẩn hóa mã SP từ Excel: float 400100000004.0 → str '400100000004'."""
        if raw is None:
            return ""
        if isinstance(raw, float):
            if raw == int(raw):
                return str(int(raw)).strip()
            return str(raw).strip()
        return str(raw).strip()

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    @staticmethod
    def generate_template():
        """Sinh Excel template bytes cho purchase.offer.line."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Các cột cần format Text để Excel không tự convert
        text_columns = {"default_code", "taxes"}

        for col_idx, (key, label, hint) in enumerate(TEMPLATE_COLUMNS, 1):
            col_letter = get_column_letter(col_idx)

            # Row 1: technical key
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

            # Row 2: label
            cell_label = ws.cell(row=2, column=col_idx, value=label)
            cell_label.font = LABEL_FONT
            cell_label.alignment = DATA_ALIGNMENT
            cell_label.border = THIN_BORDER
            if "required" in hint.lower():
                cell_label.fill = REQUIRED_FILL

            # Row 3: hint
            cell_hint = ws.cell(row=3, column=col_idx, value=hint)
            cell_hint.font = HINT_FONT
            cell_hint.fill = HINT_FILL
            cell_hint.alignment = DATA_ALIGNMENT
            cell_hint.border = THIN_BORDER

            max_len = max(len(key), len(label), len(hint))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Pre-format 1000 dòng data: font Times New Roman, căn trái, cột text format @
        text_col_indices = [
            i for i, (k, _, _) in enumerate(TEMPLATE_COLUMNS, 1) if k in text_columns
        ]
        for row in range(4, 1004):
            for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
                c = ws.cell(row=row, column=col_idx)
                c.font = Font(name=BASE_FONT_NAME, size=11)
                c.alignment = DATA_ALIGNMENT
                if col_idx in text_col_indices:
                    c.number_format = numbers.FORMAT_TEXT

        ws.freeze_panes = "A4"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
